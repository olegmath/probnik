#!/usr/bin/env python3
"""Small Soholms proxy/parser backend for the marathon rating site.

The server keeps the Soholms Authorization token on the backend side, downloads
attendance XLSX exports, parses them, and returns JSON rows compatible with the
current frontend rating calculations.
"""

from __future__ import annotations

import gzip
import hmac
import io
import json
import os
import re
import base64
import copy
import csv
import sys
import threading
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from typing import Any
from urllib.parse import parse_qs, urlencode, urlparse
from xml.sax.saxutils import escape as xml_escape

import requests
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Flowable
from reportlab.platypus import Image as ReportImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def load_env_file(path: str) -> None:
    if not os.path.exists(path):
        return
    with open(path, "r", encoding="utf-8") as file:
        for line in file:
            stripped = line.strip()
            if not stripped or stripped.startswith("#") or "=" not in stripped:
                continue
            key, value = stripped.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


load_env_file(os.path.join(os.path.dirname(__file__), ".env"))

API_BASE = os.getenv("SOHOLMS_API_BASE", "https://api.soholms.com").rstrip("/")
SOHOLMS_ORG = os.getenv("SOHOLMS_ORG", "943671").strip()
DEFAULT_CACHE_SECONDS = int(os.getenv("SOHOLMS_CACHE_SECONDS", "900"))
DEFAULT_CONCURRENCY = int(os.getenv("SOHOLMS_CONCURRENCY", "4"))
ATTEMPT_CONCURRENCY = int(os.getenv("SOHOLMS_ATTEMPT_CONCURRENCY", str(DEFAULT_CONCURRENCY * 2)))
MAX_GROUPS_PER_REQUEST = int(os.getenv("SOHOLMS_MAX_GROUPS", "80"))
DEADLINE_SHIFT_DAYS = int(os.getenv("SOHOLMS_DEADLINE_SHIFT_DAYS", "1"))
DEFAULT_PERIOD_FROM = os.getenv("SOHOLMS_DEFAULT_PERIOD_FROM", "2026-04-01")
DEFAULT_DATA_SOURCE = os.getenv("MARATHON_DATA_SOURCE", "soholms").strip().lower()
DEFAULT_CONFIG_PATH = os.path.join(os.path.dirname(__file__), "groups.config.json")
BACKEND_ADMIN_KEY = os.getenv("BACKEND_ADMIN_KEY", "").strip()
APP_SETTINGS_PATH = os.getenv(
    "APP_SETTINGS_PATH",
    os.path.join(os.path.dirname(__file__), "app_settings.json"),
)
PUBLIC_RATINGS_SNAPSHOT_PATH = os.getenv(
    "PUBLIC_RATINGS_SNAPSHOT_PATH",
    os.path.join(os.path.dirname(__file__), "public_ratings_snapshot.json"),
)
PUBLIC_RATINGS_REFRESH_SECONDS = int(os.getenv("PUBLIC_RATINGS_REFRESH_SECONDS", str(24 * 60 * 60)))
ADMIN_RATINGS_SNAPSHOT_PATH = os.getenv(
    "ADMIN_RATINGS_SNAPSHOT_PATH",
    os.path.join(os.path.dirname(__file__), "admin_ratings_snapshot.json"),
)
ADMIN_RATINGS_REFRESH_SECONDS = int(os.getenv("ADMIN_RATINGS_REFRESH_SECONDS", str(60 * 60)))
PENALTY_OVERRIDES_PATH = os.getenv(
    "PENALTY_OVERRIDES_PATH",
    os.path.join(os.path.dirname(__file__), "penalty_overrides.json"),
)
DEFAULT_TELEGRAM_CHAT_CONFIG_PATH = os.path.join(os.path.dirname(__file__), "telegram_chats.json")
TELEGRAM_API_BASE = os.getenv("TELEGRAM_API_BASE", "https://api.telegram.org").rstrip("/")
MARATHON_PLAN_CSV_PATH = os.getenv("MARATHON_PLAN_CSV_PATH", "").strip()
QUESTION_TOPICS_XLSX_PATH = os.getenv(
    "QUESTION_TOPICS_XLSX_PATH",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "question_topics.xlsx"),
).strip()

GROUP_TREE_PATH = "/api/v1/learning_group/get_tree"
ATTENDANCE_PATH = "/master/api/learning/attendance-sheet/excel/data"
GRAPHQL_PATH = "/master/graphql"

DEFAULT_ATTEMPT_DISCIPLINE_IDS = tuple(
    int(part)
    for part in os.getenv(
        "SOHOLMS_ATTEMPT_DISCIPLINE_IDS",
        "49949,49950,49947,52837,49948,49957,49956,49954,49955,49952,49951,49953",
    ).split(",")
    if part.strip()
)
FIRST_ATTEMPTS_ENABLED = os.getenv("SOHOLMS_FIRST_ATTEMPTS_ENABLED", "1").strip().lower() not in {"0", "false", "no", "off"}

DISCIPLINE_HOMEWORKS_QUERY = """
query AcademicDisciplineHomeworkIds_Query($id: ID!) {
  node(id: $id) {
    __typename
    ... on AcademicDiscipline {
      uid
      name
      treeContentFlat {
        nodes {
          id
          name
          kind
          parentNodeId
          orderIndex
          data {
            lessonContentItem {
              academicHomeworkId
            }
            interactiveLesson {
              uid
              startsAt
              finishesAt
              interactiveLessonSeriesId
            }
          }
        }
      }
    }
  }
}
""".strip()

INTERACTIVE_LESSON_HOMEWORKS_QUERY = """
query InteractiveLessonHomeworkIds_Query($id: ID!) {
  node(id: $id) {
    __typename
    ... on InteractiveLesson {
      uid
      name
      startsAt
      finishesAt
      content {
        items {
          homework {
            academicHomeworkId
          }
        }
      }
    }
  }
}
""".strip()

HOMEWORK_ATTEMPTS_QUERY = """
query AcademicHomeworkResultsStore_Query($id: ID!, $learningGroupIds: [Int!]) {
  node(id: $id) {
    __typename
    ... on AcademicHomework {
      uid
      kind
      academicDisciplineId
      learningDisciplineHomeworks(learningGroupIds: $learningGroupIds) {
        uid
        grade
        progressStatus
        deadlineAt
        academicHomeworkId
        scores
        id
        learningDiscipline {
          masterClientId
          human {
            firstName
            lastName
            middleName
          }
        }
        results {
          uid
          status
          statusChangedAt
          grade
          scores
          quiz {
            isPreventProgressWorkflow
            questionAnswers {
              uid
              scores
              isCorrect
            }
          }
        }
      }
    }
  }
}
""".strip()

SUBJECT_LABELS = {
    "математика": "Математика",
    "русский язык": "Русский язык",
    "физика": "Физика",
    "информатика": "Информатика",
    "обществознание": "Обществознание",
    "история": "История",
    "без предмета": "Без предмета",
}

SUBJECT_ALIASES = (
    ("математика", ("матем", "мат11", "м2", "м1")),
    ("русский язык", ("рус", "ря", "русский")),
    ("физика", ("физ", "физика")),
    ("информатика", ("инф", "информ")),
    ("обществознание", ("общ", "обществ")),
    ("история", ("ист", "история")),
)

MONTHS_RU = {
    1: "янв",
    2: "фев",
    3: "мар",
    4: "апр",
    5: "мая",
    6: "июн",
    7: "июл",
    8: "авг",
    9: "сен",
    10: "окт",
    11: "ноя",
    12: "дек",
}

_CACHE: dict[str, tuple[float, Any]] = {}
_PDF_FONT_NAMES: tuple[str, str] | None = None
_PENALTY_OVERRIDES_LOCK = threading.Lock()
_PUBLIC_RATINGS_SNAPSHOT_LOCK = threading.Lock()
_PUBLIC_RATINGS_REFRESH_LOCK = threading.Lock()
_PUBLIC_RATINGS_REFRESHING_LOCK = threading.Lock()
_PUBLIC_RATINGS_REFRESHING_KEYS: set[str] = set()
_ADMIN_RATINGS_SNAPSHOT_LOCK = threading.Lock()
_ADMIN_RATINGS_REFRESH_LOCK = threading.Lock()
_ADMIN_RATINGS_REFRESHING_LOCK = threading.Lock()
_ADMIN_RATINGS_REFRESHING_KEYS: set[str] = set()


class BackendError(Exception):
    def __init__(self, message: str, status: int = 500):
        super().__init__(message)
        self.status = status


class PdfIcon(Flowable):
    def __init__(self, kind: str, size: float = 7 * mm):
        super().__init__()
        self.kind = kind
        self.width = size
        self.height = size

    def draw(self) -> None:
        canvas = self.canv
        canvas.saveState()
        canvas.setFillColor(colors.HexColor("#4562f0"))
        canvas.roundRect(0, 0, self.width, self.height, 2.2, fill=1, stroke=0)
        canvas.setStrokeColor(colors.white)
        canvas.setFillColor(colors.white)
        canvas.setLineWidth(1.8)

        if self.kind == "check":
            canvas.line(self.width * 0.27, self.height * 0.52, self.width * 0.43, self.height * 0.35)
            canvas.line(self.width * 0.43, self.height * 0.35, self.width * 0.73, self.height * 0.68)
        else:
            bar_width = self.width * 0.12
            gap = self.width * 0.09
            start = self.width * 0.28
            base = self.height * 0.24
            heights = (self.height * 0.36, self.height * 0.52, self.height * 0.25)
            for index, height in enumerate(heights):
                x = start + index * (bar_width + gap)
                canvas.rect(x, base, bar_width, height, fill=1, stroke=0)

        canvas.restoreState()


@dataclass(frozen=True)
class GroupInfo:
    id: int
    name: str
    subject: str
    teacher: str
    parent_group_ids: tuple[int, ...]
    student_count: int


def clean_authorization_header(value: str) -> str:
    token = value.strip()
    token = re.sub(r"^\s*Authorization:\s*", "", token, flags=re.IGNORECASE).strip()
    token = token.strip("\"'")
    while token.endswith("\\"):
        token = token[:-1].strip()
        token = token.strip("\"'")
    token = token.strip("\"'")
    return token


def get_authorization_header(kind: str = "api") -> str:
    if kind == "excel":
        names = ("SOHOLMS_EXCEL_TOKEN", "SOHOLMS_TOKEN")
    elif kind == "graphql":
        names = ("SOHOLMS_GRAPHQL_TOKEN", "SOHOLMS_EXCEL_TOKEN", "SOHOLMS_TOKEN")
    else:
        names = ("SOHOLMS_API_TOKEN", "SOHOLMS_TOKEN")

    token = ""
    for name in names:
        token = clean_authorization_header(os.getenv(name, ""))
        if token:
            break
    if not token:
        raise BackendError(
            f"{' or '.join(names)} is not configured",
            HTTPStatus.INTERNAL_SERVER_ERROR,
        )
    return token


def soholms_headers(accept: str = "application/json", kind: str = "api") -> dict[str, str]:
    return {
        "accept": accept,
        "authorization": get_authorization_header(kind),
        "origin": "https://master.soholms.com",
        "referer": "https://master.soholms.com/",
        "user-agent": "marathon-rating-backend/1.0",
    }


def cached(key: str, ttl_seconds: int, loader):
    now = time.time()
    item = _CACHE.get(key)
    if item and now - item[0] < ttl_seconds:
        return item[1]
    try:
        value = loader()
        _CACHE[key] = (now, value)
        return value
    except Exception:
        if item:
            return item[1]
        raise


def clear_cache() -> None:
    _CACHE.clear()


def token_fingerprint(name: str) -> dict[str, Any]:
    value = os.getenv(name, "")
    stripped = clean_authorization_header(value)
    return {
        "configured": bool(stripped),
        "length": len(stripped),
        "startsWithBearer": stripped.lower().startswith("bearer "),
        "preview": f"{stripped[:10]}...{stripped[-6:]}" if len(stripped) > 20 else "",
    }


def normalize_data_source(value: Any) -> str:
    source = normalize_text(value).lower()
    return source if source in {"google", "soholms"} else ""


def default_data_source() -> str:
    return normalize_data_source(DEFAULT_DATA_SOURCE) or "soholms"


def load_app_settings() -> dict[str, Any]:
    settings = {"dataSource": default_data_source()}
    if not os.path.exists(APP_SETTINGS_PATH):
        return settings
    try:
        with open(APP_SETTINGS_PATH, "r", encoding="utf-8") as file:
            stored = json.load(file)
    except (OSError, json.JSONDecodeError):
        return settings
    if not isinstance(stored, dict):
        return settings
    data_source = normalize_data_source(stored.get("dataSource"))
    if data_source:
        settings["dataSource"] = data_source
    return settings


def save_app_settings(settings: dict[str, Any]) -> dict[str, Any]:
    data_source = normalize_data_source(settings.get("dataSource"))
    if not data_source:
        raise BackendError("dataSource must be google or soholms", HTTPStatus.BAD_REQUEST)

    payload = {
        "dataSource": data_source,
        "updatedAt": datetime.utcnow().isoformat(timespec="seconds") + "Z",
    }
    directory = os.path.dirname(APP_SETTINGS_PATH)
    if directory:
        os.makedirs(directory, exist_ok=True)
    with open(APP_SETTINGS_PATH, "w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)
        file.write("\n")
    return payload


def request_json(method: str, path: str, **kwargs) -> Any:
    url = f"{API_BASE}{path}"
    last_error: Exception | None = None
    for attempt in range(3):
        try:
            response = requests.request(method, url, timeout=45, **kwargs)
            if response.status_code >= 400:
                raise BackendError(
                    f"Soholms API error {response.status_code}: {response.text[:500]}",
                    response.status_code,
                )
            return response.json()
        except BackendError:
            raise
        except requests.RequestException as error:
            last_error = error
            if attempt < 2:
                time.sleep(1.5 * (attempt + 1))
    raise BackendError(f"Soholms API request failed: {last_error}", HTTPStatus.BAD_GATEWAY)


def to_graphql_gid(type_name: str, uid: int | str) -> str:
    raw = f"{type_name}:{uid}".encode("utf-8")
    return base64.b64encode(raw).decode("ascii")


def request_graphql(query: str, variables: dict[str, Any], query_name: str) -> dict[str, Any]:
    payload = {
        "query": query,
        "variables": variables,
    }
    data = request_json(
        "POST",
        GRAPHQL_PATH,
        headers={
            **soholms_headers(kind="graphql"),
            "content-type": "application/json",
            "x-procraft-query": query_name,
            **({"x-procraft-org": SOHOLMS_ORG} if SOHOLMS_ORG else {}),
        },
        json=payload,
    )
    errors = data.get("errors") if isinstance(data, dict) else None
    if errors:
        raise BackendError(f"Soholms GraphQL {query_name} error: {json.dumps(errors, ensure_ascii=False)[:500]}")
    return data


def parse_soholms_datetime(value: Any) -> datetime | None:
    text = normalize_text(value)
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None


def soholms_human_name(human: dict[str, Any]) -> str:
    parts = [
        normalize_text(human.get("lastName")),
        normalize_text(human.get("firstName")),
        normalize_text(human.get("middleName")),
    ]
    return " ".join(part for part in parts if part)


def normalize_person_key(value: Any) -> str:
    return normalize_text(value).casefold().replace("ё", "е")


def dedupe_ints(values: list[Any]) -> list[int]:
    seen: set[int] = set()
    result: list[int] = []
    for value in values:
        try:
            number = int(value)
        except (TypeError, ValueError):
            continue
        if number in seen:
            continue
        seen.add(number)
        result.append(number)
    return result


def day_number_from_name(value: Any) -> int | None:
    match = re.search(r"\bдень\s*(\d+)\b", normalize_text(value), flags=re.IGNORECASE)
    if not match:
        return None
    return int(match.group(1))


def lesson_date_from_schedule(starts_at_value: Any, finishes_at_value: Any) -> str:
    text = normalize_text(finishes_at_value) or normalize_text(starts_at_value)
    if not text:
        return ""
    scheduled_at = parse_soholms_datetime(text)
    return scheduled_at.date().isoformat() if scheduled_at else text[:10]


MARATHON_PLAN_CACHE: dict[str, Any] = {"path": None, "mtime": None, "items": []}


def marathon_plan_candidate_paths() -> list[str]:
    candidates = [
        MARATHON_PLAN_CSV_PATH,
        os.path.join(os.path.dirname(__file__), "marathon_plan.csv"),
        os.path.join(os.path.dirname(os.path.dirname(__file__)), "marathon_plan.csv"),
        os.path.expanduser("~/Downloads/Марафон - Лист1.csv"),
    ]
    result: list[str] = []
    for path in candidates:
        if path and path not in result:
            result.append(path)
    return result


def parse_marathon_date(value: Any) -> str:
    text = normalize_text(value)
    match = re.match(r"^(\d{1,2})[./](\d{1,2})(?:[./](\d{2,4}))?$", text)
    if not match:
        return ""
    day = int(match.group(1))
    month = int(match.group(2))
    year_text = match.group(3)
    if year_text:
        year = int(year_text)
        if year < 100:
            year += 2000
    else:
        year = datetime.fromisoformat(DEFAULT_PERIOD_FROM).year
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return ""


def expand_task_format(value: Any) -> list[dict[str, Any]]:
    text = normalize_text(value)
    if not text:
        return []
    lowered = text.casefold()
    lowered = re.sub(r"\bс\s*(\d+)\s*по\s*(\d+)\b", r"\1-\2", lowered)
    tasks: list[dict[str, Any]] = []
    for match in re.finditer(r"\d+\s*-\s*\d+|\d+", lowered):
        token = match.group(0)
        if "-" in token:
            start_text, end_text = re.split(r"\s*-\s*", token, maxsplit=1)
            start = int(start_text)
            end = int(end_text)
            step = 1 if end >= start else -1
            for number in range(start, end + step, step):
                tasks.append(
                    {
                        "key": str(number),
                        "label": str(number),
                        "number": number,
                        "source": text,
                    }
                )
        else:
            number = int(token)
            tasks.append(
                {
                    "key": str(number),
                    "label": str(number),
                    "number": number,
                    "source": text,
                }
            )
    return tasks


def task_repeat_count(count_value: Any, task_count: int) -> int:
    text = normalize_text(count_value).casefold()
    if not text or task_count <= 0:
        return 1
    per_match = re.search(r"\bпо\s*(\d+)\b", text)
    if per_match:
        return max(1, int(per_match.group(1)))
    numbers = [int(match.group(0)) for match in re.finditer(r"\d+", text)]
    if not numbers:
        return 1
    total = numbers[0]
    if task_count == 1:
        return max(1, total)
    if total > task_count and total % task_count == 0:
        return max(1, total // task_count)
    return 1


def repeat_planned_tasks(tasks: list[dict[str, Any]], count_value: Any) -> list[dict[str, Any]]:
    repeat_count = task_repeat_count(count_value, len(tasks))
    if repeat_count <= 1:
        return tasks
    repeated: list[dict[str, Any]] = []
    for task in tasks:
        for repeat_index in range(1, repeat_count + 1):
            repeated.append(
                {
                    **task,
                    "repeatIndex": repeat_index,
                    "repeatTotal": repeat_count,
                }
            )
    return repeated


def split_topic_section(topic_value: Any, section_value: Any = "") -> tuple[str, str]:
    topic = normalize_text(topic_value)
    section = normalize_text(section_value)
    if section or not topic:
        return section, topic
    match = re.match(r"^(.+?\s+\d+(?:\.\d+)*)\s+(.+)$", topic)
    if match:
        return normalize_text(match.group(1)), normalize_text(match.group(2))
    return "", topic


def marathon_plan_from_csv(path: str) -> list[dict[str, Any]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as file:
        rows = list(csv.reader(file))
    if len(rows) < 4:
        return []

    width = max(len(row) for row in rows)
    rows = [row + [""] * (width - len(row)) for row in rows]
    subject_by_col: list[str] = []
    level_by_col: list[str] = []
    current_subject = ""
    current_level = ""
    for col in range(width):
        if normalize_text(rows[0][col]):
            current_subject = infer_subject(rows[0][col])
        if normalize_text(rows[1][col]):
            current_level = normalize_text(rows[1][col])
        subject_by_col.append(current_subject)
        level_by_col.append(current_level)

    items: list[dict[str, Any]] = []
    col = 0
    while col < width:
        if normalize_text(rows[2][col]).casefold() != "№":
            col += 1
            continue
        start = col
        end = start + 1
        while end < width and normalize_text(rows[2][end]):
            end += 1
        headers = {normalize_text(rows[2][index]).casefold(): index for index in range(start, end)}
        subject = subject_by_col[start]
        level = level_by_col[start]
        theme_col = headers.get("тема")
        format_col = headers.get("формат")
        section_col = headers.get("раздел")
        day_col = headers.get("№")
        date_col = headers.get("дата")
        count_col = headers.get("кол-во заданий")
        current_day_number = None
        current_date_key = ""

        for row in rows[3:]:
            row_day_number = None
            try:
                row_day_number = int(float(normalize_text(row[day_col]))) if day_col is not None and normalize_text(row[day_col]) else None
            except (TypeError, ValueError):
                row_day_number = None
            if row_day_number is not None:
                current_day_number = row_day_number
            row_date_key = parse_marathon_date(row[date_col]) if date_col is not None else ""
            if row_date_key:
                current_date_key = row_date_key
            day_number = row_day_number if row_day_number is not None else current_day_number
            topic_source = row[theme_col] if theme_col is not None else ""
            format_text = normalize_text(row[format_col]) if format_col is not None else ""
            if row_day_number is None and not normalize_text(topic_source) and not format_text:
                continue
            if day_number is None and not normalize_text(topic_source) and not format_text:
                continue
            section, topic = split_topic_section(topic_source, row[section_col] if section_col is not None else "")
            task_source = format_text if format_col is not None else topic
            task_count_text = normalize_text(row[count_col]) if count_col is not None else ""
            tasks = repeat_planned_tasks(expand_task_format(task_source), task_count_text)
            item = {
                "subject": subject,
                "level": level,
                "dayNumber": day_number,
                "dateKey": row_date_key or current_date_key,
                "section": section,
                "topic": topic,
                "format": format_text,
                "taskCount": task_count_text,
                "plannedTasks": tasks,
            }
            if item["dayNumber"] is not None or item["topic"] or item["plannedTasks"]:
                items.append(item)
        col = end + 1
    return items


def load_marathon_plan_items() -> list[dict[str, Any]]:
    for path in marathon_plan_candidate_paths():
        if not os.path.exists(path):
            continue
        try:
            mtime = os.path.getmtime(path)
            if MARATHON_PLAN_CACHE["path"] == path and MARATHON_PLAN_CACHE["mtime"] == mtime:
                return MARATHON_PLAN_CACHE["items"]
            items = marathon_plan_from_csv(path)
            MARATHON_PLAN_CACHE.update({"path": path, "mtime": mtime, "items": items})
            return items
        except Exception as error:
            sys.stderr.write(f"Failed to load marathon plan CSV {path}: {error}\n")
    MARATHON_PLAN_CACHE.update({"path": None, "mtime": None, "items": []})
    return []


def combine_marathon_plan_items(items: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not items:
        return None
    base = dict(items[0])
    for key in ("section", "topic", "format", "taskCount"):
        values: list[str] = []
        for item in items:
            value = normalize_text(item.get(key))
            if value and value not in values:
                values.append(value)
        base[key] = "; ".join(values)
    planned_tasks: list[dict[str, Any]] = []
    for item in items:
        planned_tasks.extend(item.get("plannedTasks") or [])
    base["plannedTasks"] = planned_tasks
    return base


def find_marathon_plan_item(row: dict[str, Any]) -> dict[str, Any] | None:
    subject = normalize_text(row.get("subject"))
    level = normalize_text(row.get("level"))
    day_number = int(row.get("dayNumber") or 0)
    lesson_date = normalize_text(row.get("dateKey") or row.get("lessonDate"))
    candidates = [
        item for item in load_marathon_plan_items()
        if item.get("subject") == subject and item.get("level") == level
    ]
    if day_number:
        by_day = [item for item in candidates if int(item.get("dayNumber") or 0) == day_number]
        if by_day:
            return combine_marathon_plan_items(by_day)
    if lesson_date:
        by_date = [item for item in candidates if item.get("dateKey") == lesson_date]
        if by_date:
            return combine_marathon_plan_items(by_date)
    return None


QUESTION_TOPICS_CACHE: dict[str, Any] = {"path": None, "mtime": None, "items": {}}

QUESTION_TOPICS_SUBJECT_MAP = (
    ("матем", "математика"),
    ("мат база", "математика"),
    ("мат огэ", "математика"),
    ("мат егэ", "математика"),
    ("мат ", "математика"),
    ("ря ", "русский язык"),
    ("рус", "русский язык"),
    ("общ", "обществознание"),
    ("ист", "история"),
    ("инф", "информатика"),
    ("физ", "физика"),
)


def parse_topic_section_header(header: str) -> tuple[str, str]:
    text = header.strip().lower()
    if not text:
        return "", ""
    subject = ""
    for prefix, sub in QUESTION_TOPICS_SUBJECT_MAP:
        if text.startswith(prefix):
            subject = sub
            break
    if not subject:
        return "", ""
    if "огэ" in text:
        level = "ОГЭ"
    elif "егэ" in text or "база" in text:
        level = "ЕГЭ"
    else:
        return subject, ""
    return subject, level


def load_question_topics() -> dict[tuple[str, str, int], str]:
    path = QUESTION_TOPICS_XLSX_PATH
    if not path or not os.path.exists(path):
        return {}
    try:
        mtime = os.path.getmtime(path)
        if QUESTION_TOPICS_CACHE["path"] == path and QUESTION_TOPICS_CACHE["mtime"] == mtime:
            return QUESTION_TOPICS_CACHE["items"]
        wb = load_workbook(path, data_only=True, read_only=True)
        topics: dict[tuple[str, str, int], str] = {}
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            current_subject = ""
            current_level = ""
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                first = row[0]
                second = row[1] if len(row) > 1 else None
                if isinstance(first, str) and first.strip():
                    current_subject, current_level = parse_topic_section_header(first)
                    continue
                if first is None:
                    continue
                try:
                    number = int(float(first))
                except (TypeError, ValueError):
                    continue
                topic = normalize_text(second)
                if not topic or not current_subject or not current_level:
                    continue
                key = (current_subject, current_level, number)
                if key not in topics:
                    topics[key] = topic
        QUESTION_TOPICS_CACHE.update({"path": path, "mtime": mtime, "items": topics})
        return topics
    except Exception as error:
        sys.stderr.write(f"Failed to load question topics XLSX {path}: {error}\n")
    QUESTION_TOPICS_CACHE.update({"path": None, "mtime": None, "items": {}})
    return {}


def lookup_question_topic(subject: str, level: str, exam_number: Any) -> str:
    if not subject or not level or exam_number in (None, ""):
        return ""
    try:
        number = int(float(str(exam_number).strip()))
    except (TypeError, ValueError):
        return ""
    return load_question_topics().get((subject, level, number), "")


def enrich_questions_with_marathon_plan(row: dict[str, Any]) -> dict[str, Any]:
    plan = find_marathon_plan_item(row)
    subject = normalize_text(row.get("subject"))
    level = normalize_text(row.get("level"))
    if not plan:
        questions = []
        for question in row.get("questions") or []:
            questions.append({**question, "plannedTask": question.get("plannedTask")})
        return {**row, "questions": questions}
    planned_tasks = plan.get("plannedTasks") or []
    enriched_questions = []
    for question in row.get("questions") or []:
        question_number = int(question.get("number") or 0)
        planned_task = planned_tasks[question_number - 1] if 0 < question_number <= len(planned_tasks) else None
        if planned_task:
            exam_topic = lookup_question_topic(subject, level, planned_task.get("label"))
            planned_task = {
                **planned_task,
                "section": plan.get("section") or "",
                "topic": exam_topic or plan.get("topic") or "",
            }
        enriched_questions.append({**question, "plannedTask": planned_task})
    return {
        **row,
        "marathonPlan": {
            "section": plan.get("section") or "",
            "topic": plan.get("topic") or "",
            "format": plan.get("format") or "",
            "taskCount": plan.get("taskCount") or "",
            "plannedTasks": planned_tasks,
        },
        "questions": enriched_questions,
    }


def homework_metadata_from_lesson(
    academic_homework_id: Any,
    *,
    discipline_id: Any = None,
    discipline_name: Any = "",
    subject: Any = "",
    level: Any = "",
    interactive_lesson_id: Any = None,
    day_title: Any = "",
    starts_at: Any = "",
    finishes_at: Any = "",
    order_index: Any = None,
    tree_node_id: Any = "",
    parent_node_id: Any = "",
) -> dict[str, Any] | None:
    ids = dedupe_ints([academic_homework_id])
    if not ids:
        return None
    day_title_text = normalize_text(day_title)
    day_number = day_number_from_name(day_title_text)
    return {
        "academicHomeworkId": ids[0],
        "academicDisciplineId": discipline_id,
        "disciplineName": normalize_text(discipline_name),
        "subject": normalize_text(subject),
        "level": normalize_text(level),
        "interactiveLessonId": interactive_lesson_id,
        "dayTitle": day_title_text,
        "dayNumber": day_number,
        "lessonDate": lesson_date_from_schedule(starts_at, finishes_at),
        "startsAt": normalize_text(starts_at),
        "finishesAt": normalize_text(finishes_at),
        "orderIndex": order_index,
        "treeNodeId": normalize_text(tree_node_id),
        "parentNodeId": normalize_text(parent_node_id),
    }


def fetch_interactive_lesson_homework_items(interactive_lesson_id: int, parent_metadata: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    try:
        lesson_id = int(interactive_lesson_id)
    except (TypeError, ValueError):
        return []

    def load():
        data = request_graphql(
            INTERACTIVE_LESSON_HOMEWORKS_QUERY,
            {"id": to_graphql_gid("InteractiveLesson", lesson_id)},
            "InteractiveLessonHomeworkIds_Query",
        )
        node = (data.get("data") or {}).get("node") or {}
        base_metadata = {
            **(parent_metadata or {}),
            "interactiveLessonId": node.get("uid") or lesson_id,
            "dayTitle": normalize_text(node.get("name")) or normalize_text((parent_metadata or {}).get("dayTitle")),
            "startsAt": node.get("startsAt") or (parent_metadata or {}).get("startsAt"),
            "finishesAt": node.get("finishesAt") or (parent_metadata or {}).get("finishesAt"),
        }
        rows: list[dict[str, Any]] = []
        for content_item in (((node.get("content") or {}).get("items")) or []):
            homework = content_item.get("homework") or {}
            metadata = homework_metadata_from_lesson(
                homework.get("academicHomeworkId"),
                discipline_id=base_metadata.get("academicDisciplineId"),
                discipline_name=base_metadata.get("disciplineName"),
                subject=base_metadata.get("subject"),
                level=base_metadata.get("level"),
                interactive_lesson_id=base_metadata.get("interactiveLessonId"),
                day_title=base_metadata.get("dayTitle"),
                starts_at=base_metadata.get("startsAt"),
                finishes_at=base_metadata.get("finishesAt"),
                order_index=base_metadata.get("orderIndex"),
                tree_node_id=base_metadata.get("treeNodeId"),
                parent_node_id=base_metadata.get("parentNodeId"),
            )
            if metadata:
                rows.append(metadata)
        return rows

    return cached(f"interactive_lesson_homework_items:{lesson_id}", DEFAULT_CACHE_SECONDS, load)


def fetch_interactive_lesson_homework_ids(interactive_lesson_id: int) -> list[int]:
    return dedupe_ints([item.get("academicHomeworkId") for item in fetch_interactive_lesson_homework_items(interactive_lesson_id)])


def fetch_discipline_homework_items(academic_discipline_id: int) -> list[dict[str, Any]]:
    def load():
        data = request_graphql(
            DISCIPLINE_HOMEWORKS_QUERY,
            {"id": to_graphql_gid("AcademicDiscipline", academic_discipline_id)},
            "AcademicDisciplineHomeworkIds_Query",
        )
        node = (data.get("data") or {}).get("node") or {}
        discipline_name = normalize_text(node.get("name"))
        discipline_subject = infer_subject(discipline_name)
        discipline_level = infer_level(discipline_name)
        rows: list[dict[str, Any]] = []
        for item in ((node.get("treeContentFlat") or {}).get("nodes") or []):
            data_item = item.get("data") or {}
            lesson_item = data_item.get("lessonContentItem") or {}
            interactive_lesson = data_item.get("interactiveLesson") or {}
            metadata = {
                "academicDisciplineId": node.get("uid") or academic_discipline_id,
                "disciplineName": discipline_name,
                "subject": discipline_subject,
                "level": discipline_level,
                "interactiveLessonId": interactive_lesson.get("uid"),
                "dayTitle": item.get("name"),
                "startsAt": interactive_lesson.get("startsAt"),
                "finishesAt": interactive_lesson.get("finishesAt"),
                "orderIndex": item.get("orderIndex"),
                "treeNodeId": item.get("id"),
                "parentNodeId": item.get("parentNodeId"),
            }
            direct_homework = homework_metadata_from_lesson(
                lesson_item.get("academicHomeworkId"),
                discipline_id=metadata.get("academicDisciplineId"),
                discipline_name=metadata.get("disciplineName"),
                subject=metadata.get("subject"),
                level=metadata.get("level"),
                interactive_lesson_id=metadata.get("interactiveLessonId"),
                day_title=metadata.get("dayTitle"),
                starts_at=metadata.get("startsAt"),
                finishes_at=metadata.get("finishesAt"),
                order_index=metadata.get("orderIndex"),
                tree_node_id=metadata.get("treeNodeId"),
                parent_node_id=metadata.get("parentNodeId"),
            )
            if direct_homework:
                rows.append(direct_homework)
            interactive_lesson = data_item.get("interactiveLesson") or {}
            rows.extend(fetch_interactive_lesson_homework_items(interactive_lesson.get("uid"), metadata))
        return rows

    return cached(f"discipline_homework_items:{academic_discipline_id}", DEFAULT_CACHE_SECONDS, load)


def fetch_discipline_homework_ids(academic_discipline_id: int) -> list[int]:
    return dedupe_ints([item.get("academicHomeworkId") for item in fetch_discipline_homework_items(academic_discipline_id)])


def fetch_homework_first_attempts(academic_homework_id: int) -> list[dict[str, Any]]:
    def load():
        data = request_graphql(
            HOMEWORK_ATTEMPTS_QUERY,
            {
                "id": to_graphql_gid("AcademicHomework", academic_homework_id),
                "learningGroupIds": [],
            },
            "AcademicHomeworkResultsStore_Query",
        )
        node = (data.get("data") or {}).get("node") or {}
        rows: list[dict[str, Any]] = []
        for homework in node.get("learningDisciplineHomeworks") or []:
            learning_discipline = homework.get("learningDiscipline") or {}
            master_client_id = learning_discipline.get("masterClientId")
            student_name = soholms_human_name(learning_discipline.get("human") or {})
            deadline_at = parse_soholms_datetime(homework.get("deadlineAt"))
            results = [
                result
                for result in homework.get("results", [])
                if parse_soholms_datetime(result.get("statusChangedAt"))
            ]
            if not master_client_id or not results:
                continue
            first_result = min(results, key=lambda result: parse_soholms_datetime(result.get("statusChangedAt")) or datetime.max)
            first_attempt_at = parse_soholms_datetime(first_result.get("statusChangedAt"))
            if not first_attempt_at:
                continue
            rows.append(
                {
                    "academicHomeworkId": academic_homework_id,
                    "academicDisciplineId": node.get("academicDisciplineId"),
                    "masterClientId": int(master_client_id),
                    "studentName": student_name,
                    "deadlineAt": deadline_at,
                    "firstAttemptAt": first_attempt_at,
                }
            )
        return rows
    return cached(f"homework_first_attempts:{academic_homework_id}", DEFAULT_CACHE_SECONDS, load)


def fetch_homework_error_maps(academic_homework_id: int) -> list[dict[str, Any]]:
    def load():
        data = request_graphql(
            HOMEWORK_ATTEMPTS_QUERY,
            {
                "id": to_graphql_gid("AcademicHomework", academic_homework_id),
                "learningGroupIds": [],
            },
            "AcademicHomeworkResultsStore_Query",
        )
        node = (data.get("data") or {}).get("node") or {}
        rows: list[dict[str, Any]] = []
        for homework in node.get("learningDisciplineHomeworks") or []:
            learning_discipline = homework.get("learningDiscipline") or {}
            master_client_id = learning_discipline.get("masterClientId")
            student_name = soholms_human_name(learning_discipline.get("human") or {})
            attempts = []
            for result in homework.get("results", []) or []:
                changed_at = parse_soholms_datetime(result.get("statusChangedAt"))
                answers = ((result.get("quiz") or {}).get("questionAnswers")) or []
                if not changed_at or not answers:
                    continue
                attempts.append(
                    {
                        "uid": result.get("uid"),
                        "status": result.get("status"),
                        "statusChangedAt": changed_at.isoformat(),
                        "grade": result.get("grade"),
                        "scores": result.get("scores"),
                        "answers": answers,
                    }
                )
            attempts.sort(key=lambda attempt: parse_soholms_datetime(attempt.get("statusChangedAt")) or datetime.max)
            if not master_client_id or not student_name or not attempts:
                continue

            question_order: list[str] = []
            questions: dict[str, dict[str, Any]] = {}
            for attempt_index, attempt in enumerate(attempts, start=1):
                for answer in attempt["answers"]:
                    question_uid = normalize_text(answer.get("uid"))
                    if not question_uid:
                        continue
                    if question_uid not in questions:
                        question_order.append(question_uid)
                        questions[question_uid] = {
                            "questionUid": question_uid,
                            "number": len(question_order),
                            "attempts": [],
                        }
                    questions[question_uid]["attempts"].append(
                        {
                            "attemptNumber": attempt_index,
                            "resultUid": attempt.get("uid"),
                            "statusChangedAt": attempt.get("statusChangedAt"),
                            "scores": answer.get("scores"),
                            "isCorrect": bool(answer.get("isCorrect")),
                        }
                    )

            question_rows = []
            fixed = 0
            remaining = 0
            wrong_total = 0
            for question_uid in question_order:
                question = questions[question_uid]
                wrong_attempts = [attempt for attempt in question["attempts"] if not attempt["isCorrect"]]
                if not wrong_attempts:
                    continue
                last_attempt = question["attempts"][-1]
                status = "fixed" if last_attempt["isCorrect"] else "remaining"
                if status == "fixed":
                    fixed += 1
                else:
                    remaining += 1
                wrong_total += 1
                question_rows.append(
                    {
                        **question,
                        "status": status,
                        "wrongAttempts": len(wrong_attempts),
                    }
                )

            rows.append(
                {
                    "academicHomeworkId": int(academic_homework_id),
                    "academicDisciplineId": node.get("academicDisciplineId"),
                    "homeworkUid": homework.get("uid"),
                    "studentHomeworkId": homework.get("id"),
                    "masterClientId": int(master_client_id),
                    "studentName": student_name,
                    "deadlineAt": homework.get("deadlineAt"),
                    "deadlineDate": normalize_text(homework.get("deadlineAt"))[:10],
                    "progressStatus": homework.get("progressStatus"),
                    "grade": homework.get("grade"),
                    "scores": homework.get("scores"),
                    "attempts": [
                        {
                            key: attempt.get(key)
                            for key in ("uid", "status", "statusChangedAt", "grade", "scores")
                        }
                        for attempt in attempts
                    ],
                    "summary": {
                        "wrongTotal": wrong_total,
                        "fixed": fixed,
                        "remaining": remaining,
                        "attempts": len(attempts),
                    },
                    "questions": question_rows,
                }
            )
        return rows

    return cached(f"homework_error_maps:{academic_homework_id}", DEFAULT_CACHE_SECONDS, load)


def build_error_map_payload(query: dict[str, str]) -> dict[str, Any]:
    student_query = normalize_person_key(query.get("studentName", ""))
    homework_id = normalize_text(query.get("academicHomeworkId") or query.get("homeworkId"))
    requested_subject = infer_subject(query.get("subject", ""))
    requested_level = normalize_text(query.get("level"))
    period_from = normalize_text(query.get("periodFrom"))
    period_to = normalize_text(query.get("periodTo"))
    if not student_query:
        raise BackendError("studentName is required", HTTPStatus.BAD_REQUEST)

    if homework_id:
        homework_ids = dedupe_ints([homework_id])
        homework_metadata: dict[int, dict[str, Any]] = {}
    else:
        discipline_ids = DEFAULT_ATTEMPT_DISCIPLINE_IDS
        with ThreadPoolExecutor(max_workers=DEFAULT_CONCURRENCY) as executor:
            results = list(executor.map(fetch_discipline_homework_items, discipline_ids))
        homework_items = [item for items in results for item in items]
        if requested_subject != "без предмета":
            homework_items = [item for item in homework_items if item.get("subject") == requested_subject]
        if requested_level:
            homework_items = [item for item in homework_items if item.get("level") == requested_level]
        homework_ids = dedupe_ints(item.get("academicHomeworkId") for item in homework_items)
        homework_metadata = {
            int(item["academicHomeworkId"]): item
            for item in homework_items
            if item.get("academicHomeworkId")
        }

    unfinished_progress_statuses = {"InProgress", "NotStarted", "Started", "Draft", ""}

    maps: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    seen_hw_ids: set[int] = set()
    with ThreadPoolExecutor(max_workers=ATTEMPT_CONCURRENCY) as executor:
        futures = {executor.submit(fetch_homework_error_maps, hw_id): hw_id for hw_id in homework_ids}
        for future in as_completed(futures):
            hw_id = futures[future]
            try:
                for row in future.result():
                    if student_query in normalize_person_key(row.get("studentName")):
                        merged_row = {
                            **row,
                            **homework_metadata.get(int(row.get("academicHomeworkId") or 0), {}),
                        }
                        display_date = error_map_display_date(merged_row)
                        if not date_in_period(display_date, period_from, period_to):
                            continue
                        progress_status = normalize_text(merged_row.get("progressStatus"))
                        is_completed = progress_status not in unfinished_progress_statuses
                        merged_row["dateKey"] = display_date
                        merged_row["completed"] = is_completed
                        merged_row = enrich_questions_with_marathon_plan(merged_row)
                        maps.append(merged_row)
                        seen_hw_ids.add(int(row.get("academicHomeworkId") or 0))
            except Exception as error:
                errors.append({"academicHomeworkId": hw_id, "error": str(error)})

    today_iso = datetime.now().date().isoformat()
    for hw_id in homework_ids:
        if int(hw_id) in seen_hw_ids:
            continue
        meta = homework_metadata.get(int(hw_id))
        if not meta:
            continue
        lesson_date = normalize_text(meta.get("lessonDate"))[:10]
        if lesson_date and lesson_date > today_iso:
            continue
        synthetic_row: dict[str, Any] = {
            **meta,
            "academicHomeworkId": int(hw_id),
            "studentName": query.get("studentName", ""),
            "completed": False,
            "questions": [],
            "summary": {"wrongTotal": 0, "fixed": 0, "remaining": 0, "attempts": 0},
        }
        display_date = error_map_display_date(synthetic_row)
        if not date_in_period(display_date, period_from, period_to):
            continue
        synthetic_row["dateKey"] = display_date
        synthetic_row = enrich_questions_with_marathon_plan(synthetic_row)
        maps.append(synthetic_row)

    maps.sort(key=lambda row: (
        int(row.get("dayNumber") or 0),
        int(row.get("orderIndex") or 0),
        normalize_text(row.get("lessonDate") or row.get("deadlineAt")),
        int(row.get("academicHomeworkId") or 0),
    ))

    day_groups: list[dict[str, Any]] = []
    day_index: dict[tuple, int] = {}
    for row in maps:
        day_key = (
            normalize_text(row.get("dayTitle") or ""),
            int(row.get("dayNumber") or 0),
            int(row.get("orderIndex") or 0),
        )
        if day_key not in day_index:
            day_index[day_key] = len(day_groups)
            day_groups.append({
                "dayTitle": row.get("dayTitle") or "",
                "dayNumber": row.get("dayNumber"),
                "dayOrder": row.get("dayNumber"),
                "orderIndex": row.get("orderIndex"),
                "dateKey": row.get("dateKey") or "",
                "marathonPlan": row.get("marathonPlan"),
                "completed": False,
                "maps": [],
            })
        if row.get("completed"):
            day_groups[day_index[day_key]]["completed"] = True
        day_groups[day_index[day_key]]["maps"].append({
            "questions": row.get("questions") or [],
            "summary": row.get("summary") or {},
            "marathonPlan": row.get("marathonPlan"),
            "completed": bool(row.get("completed")),
        })

    total_days = len(day_groups)
    completed_days = sum(1 for d in day_groups if d.get("completed"))
    unfinished_days = total_days - completed_days
    total_errors = sum(
        int((m.get("summary") or {}).get("wrongTotal") or 0)
        for d in day_groups for m in (d.get("maps") or [])
    )
    total_fixed = sum(
        int((m.get("summary") or {}).get("fixed") or 0)
        for d in day_groups for m in (d.get("maps") or [])
    )
    total_remaining = sum(
        int((m.get("summary") or {}).get("remaining") or 0)
        for d in day_groups for m in (d.get("maps") or [])
    )

    return {
        "ok": True,
        "studentName": query.get("studentName", ""),
        "period": {"from": period_from, "to": period_to},
        "homeworkCount": len(homework_ids),
        "summary": {
            "totalDays": total_days,
            "completedDays": completed_days,
            "unfinishedDays": unfinished_days,
            "totalErrors": total_errors,
            "fixed": total_fixed,
            "remaining": total_remaining,
        },
        "maps": day_groups,
        "errors": errors,
    }


def error_map_display_date(row: dict[str, Any]) -> str:
    lesson_date = normalize_text(row.get("lessonDate"))
    if lesson_date:
        return shift_iso_date(lesson_date, DEADLINE_SHIFT_DAYS)
    return normalize_text(row.get("deadlineDate")) or normalize_text(row.get("deadlineAt"))[:10]


def date_in_period(date_key: str, period_from: str, period_to: str) -> bool:
    if not date_key:
        return True
    if period_from and date_key < period_from:
        return False
    if period_to and date_key > period_to:
        return False
    return True


def error_homework_items_for_query(query: dict[str, str]) -> tuple[list[int], dict[int, dict[str, Any]]]:
    requested_subject = infer_subject(query.get("subject", ""))
    requested_level = normalize_text(query.get("level"))
    discipline_ids = DEFAULT_ATTEMPT_DISCIPLINE_IDS
    with ThreadPoolExecutor(max_workers=DEFAULT_CONCURRENCY) as executor:
        results = list(executor.map(fetch_discipline_homework_items, discipline_ids))
    homework_items = [item for items in results for item in items]
    if requested_subject != "без предмета":
        homework_items = [item for item in homework_items if item.get("subject") == requested_subject]
    if requested_level:
        homework_items = [item for item in homework_items if item.get("level") == requested_level]
    homework_ids = dedupe_ints(item.get("academicHomeworkId") for item in homework_items)
    homework_metadata = {
        int(item["academicHomeworkId"]): item
        for item in homework_items
        if item.get("academicHomeworkId")
    }
    return homework_ids, homework_metadata


def selected_error_analytics_students(query: dict[str, str]) -> list[dict[str, Any]]:
    payload = resolve_ratings_payload(query)
    requested_subject = infer_subject(query.get("subject", ""))
    requested_level = normalize_text(query.get("level"))
    requested_group = normalize_text(query.get("group"))
    requested_teacher = normalize_text(query.get("teacher"))
    rows = []
    seen: set[str] = set()
    for row in payload.get("rows", []) or []:
        if requested_subject != "без предмета" and row.get("subject") != requested_subject:
            continue
        if requested_level and row.get("level") != requested_level:
            continue
        if requested_group and row.get("group") != requested_group:
            continue
        if requested_teacher and row.get("teacher") != requested_teacher:
            continue
        key = normalize_person_key(row.get("name"))
        if not key or key in seen:
            continue
        seen.add(key)
        rows.append(row)
    return rows


def build_error_analytics_payload(query: dict[str, str]) -> dict[str, Any]:
    selected_students = selected_error_analytics_students(query)
    student_keys = {normalize_person_key(row.get("name")) for row in selected_students}
    student_names = {
        normalize_person_key(row.get("name")): normalize_text(row.get("name"))
        for row in selected_students
    }
    period_from = normalize_text(query.get("periodFrom"))
    period_to = normalize_text(query.get("periodTo"))
    homework_ids, homework_metadata = error_homework_items_for_query(query)

    days: dict[str, dict[str, Any]] = {}
    student_summaries: dict[str, dict[str, Any]] = {
        key: {
            "studentName": student_names.get(key, ""),
            "wrongTotal": 0,
            "fixed": 0,
            "remaining": 0,
        }
        for key in student_keys
    }
    errors: list[dict[str, Any]] = []

    with ThreadPoolExecutor(max_workers=ATTEMPT_CONCURRENCY) as executor:
        futures = {executor.submit(fetch_homework_error_maps, hw_id): hw_id for hw_id in homework_ids}
        for future in as_completed(futures):
            hw_id = futures[future]
            metadata = homework_metadata.get(int(hw_id), {})
            try:
                homework_rows = future.result()
            except Exception as error:
                errors.append({"academicHomeworkId": hw_id, "error": str(error)})
                continue

            for row in homework_rows:
                student_key = normalize_person_key(row.get("studentName"))
                if student_key not in student_keys:
                    continue
                merged_row = {**row, **metadata}
                display_date = error_map_display_date(merged_row)
                if not date_in_period(display_date, period_from, period_to):
                    continue
                merged_row["dateKey"] = display_date
                merged_row = enrich_questions_with_marathon_plan(merged_row)
                day_key = normalize_text(merged_row.get("dayTitle")) or normalize_text(merged_row.get("lessonDate")) or str(hw_id)
                day = days.setdefault(
                    day_key,
                    {
                        "dayTitle": merged_row.get("dayTitle") or "",
                        "dayNumber": merged_row.get("dayNumber"),
                        "lessonDate": display_date or merged_row.get("lessonDate") or "",
                        "orderIndex": merged_row.get("orderIndex"),
                        "marathonPlan": merged_row.get("marathonPlan") or {},
                        "wrongTotal": 0,
                        "fixed": 0,
                        "remaining": 0,
                        "questions": {},
                    },
                )
                student_summary = student_summaries.setdefault(
                    student_key,
                    {
                        "studentName": student_names.get(student_key) or merged_row.get("studentName") or "",
                        "wrongTotal": 0,
                        "fixed": 0,
                        "remaining": 0,
                    },
                )
                for question in merged_row.get("questions") or []:
                    question_number = int(question.get("number") or 0)
                    if not question_number:
                        continue
                    planned_task = question.get("plannedTask") or {}
                    day_plan = merged_row.get("marathonPlan") or {}
                    question_key = normalize_text(planned_task.get("key")) or str(question_number)
                    task_label = normalize_text(planned_task.get("label")) or str(question_number)
                    status = question.get("status")
                    wrong_attempts = int(question.get("wrongAttempts") or 0)
                    q = day["questions"].setdefault(
                        question_key,
                        {
                            "number": question_number,
                            "taskKey": question_key,
                            "taskLabel": task_label,
                            "taskNumber": planned_task.get("number"),
                            "section": planned_task.get("section") or day_plan.get("section") or "",
                            "topic": planned_task.get("topic") or day_plan.get("topic") or "",
                            "studentsWrong": 0,
                            "studentsFixed": 0,
                            "studentsRemaining": 0,
                            "wrongAttempts": 0,
                            "remainingStudents": [],
                            "fixedStudents": [],
                        },
                    )
                    q["studentsWrong"] += 1
                    q["wrongAttempts"] += wrong_attempts
                    day["wrongTotal"] += 1
                    student_summary["wrongTotal"] += 1
                    if status == "fixed":
                        q["studentsFixed"] += 1
                        q["fixedStudents"].append(student_summary["studentName"])
                        day["fixed"] += 1
                        student_summary["fixed"] += 1
                    else:
                        q["studentsRemaining"] += 1
                        q["remainingStudents"].append(student_summary["studentName"])
                        day["remaining"] += 1
                        student_summary["remaining"] += 1

    day_rows = []
    for day in days.values():
        questions = sorted(
            day.pop("questions").values(),
            key=lambda item: (
                int(item.get("taskNumber") or item.get("number") or 0),
                normalize_text(item.get("taskLabel")),
            ),
        )
        day_rows.append({**day, "questions": questions})
    day_rows.sort(key=lambda day: (
        int(day.get("dayNumber") or 0),
        int(day.get("orderIndex") or 0),
        normalize_text(day.get("lessonDate")),
    ))
    student_rows = sorted(
        student_summaries.values(),
        key=lambda item: (-int(item.get("remaining") or 0), -int(item.get("wrongTotal") or 0), item.get("studentName") or ""),
    )
    summary = {
        "students": len(selected_students),
        "studentsWithErrors": sum(1 for item in student_rows if int(item.get("wrongTotal") or 0) > 0),
        "wrongTotal": sum(int(day.get("wrongTotal") or 0) for day in day_rows),
        "fixed": sum(int(day.get("fixed") or 0) for day in day_rows),
        "remaining": sum(int(day.get("remaining") or 0) for day in day_rows),
    }
    return {
        "ok": True,
        "summary": summary,
        "days": day_rows,
        "students": student_rows,
        "homeworkCount": len(homework_ids),
        "errors": errors,
    }


def build_first_attempt_index(period_from: str, period_to: str) -> dict[tuple[str, str], datetime]:
    if not FIRST_ATTEMPTS_ENABLED:
        return {}

    def load():
        t_total = time.time()
        period_start = datetime.fromisoformat(period_from).date() if period_from else None
        period_end = datetime.fromisoformat(period_to).date() if period_to else None

        # Получаем метаданные ДЗ (subject, level, dayNumber из названия "День N")
        t0 = time.time()
        all_items: list[dict[str, Any]] = []
        with ThreadPoolExecutor(max_workers=DEFAULT_CONCURRENCY) as executor:
            results = list(executor.map(fetch_discipline_homework_items, DEFAULT_ATTEMPT_DISCIPLINE_IDS))
        for items in results:
            all_items.extend(items)

        # hw_id → (дата дедлайна, subject, level) — без subject/level одна и та же дата у разных предметов схлопывается в один ключ
        hw_plan_meta: dict[int, tuple[date, str, str]] = {}
        for item in all_items:
            hw_id = item.get("academicHomeworkId")
            if not hw_id:
                continue
            plan = find_marathon_plan_item(item)
            date_str = plan.get("dateKey") if plan else None
            if date_str:
                try:
                    plan_date = datetime.fromisoformat(date_str[:10]).date()
                except (ValueError, TypeError):
                    continue
                subject = normalize_text(item.get("subject")).casefold()
                level = normalize_text(item.get("level")).casefold()
                hw_plan_meta[hw_id] = (plan_date, subject, level)

        homework_ids = list(hw_plan_meta.keys())
        t_disciplines = time.time() - t0

        t1 = time.time()
        index: dict[tuple[str, str], tuple[datetime, date]] = {}
        with ThreadPoolExecutor(max_workers=ATTEMPT_CONCURRENCY) as executor:
            futures = {executor.submit(fetch_homework_first_attempts, hw_id): hw_id for hw_id in homework_ids}
            for future in as_completed(futures):
                hw_id = futures[future]
                meta = hw_plan_meta.get(hw_id)
                if not meta:
                    continue
                plan_date, subject, level = meta
                if period_start and plan_date < period_start:
                    continue
                if period_end and plan_date > period_end:
                    continue
                scope = f"{subject}|{level}"
                for row in future.result():
                    keys = [(f"id:{row['masterClientId']}|{scope}", plan_date.isoformat())]
                    student_name = normalize_person_key(row.get("studentName"))
                    if student_name:
                        keys.append((f"name:{student_name}|{scope}", plan_date.isoformat()))
                    for key in keys:
                        previous = index.get(key)
                        if previous is None or row["firstAttemptAt"] < previous[0]:
                            index[key] = (row["firstAttemptAt"], plan_date)
        t_homeworks = time.time() - t1

        print(f"[timings] build_first_attempt_index: disciplines={t_disciplines:.2f}s homeworks={t_homeworks:.2f}s total={time.time()-t_total:.2f}s n_homeworks={len(homework_ids)}", flush=True, file=sys.stderr)
        return index

    cache_key = f"first_attempt_index:{period_from}:{period_to}:{DEFAULT_ATTEMPT_DISCIPLINE_IDS}"
    return cached(cache_key, DEFAULT_CACHE_SECONDS, load)


def fetch_group_tree() -> list[dict[str, Any]]:
    def load():
        data = request_json(
            "POST",
            GROUP_TREE_PATH,
            headers={**soholms_headers(), "content-type": "application/json"},
            data="",
        )
        groups = data.get("learningGroups")
        if not isinstance(groups, list):
            raise BackendError("Unexpected get_tree response: missing learningGroups")
        return groups

    return cached("group_tree", DEFAULT_CACHE_SECONDS, load)


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_group_name(value: Any) -> str:
    return normalize_text(value).casefold()


def compact_group_name(value: Any) -> str:
    return re.sub(r"[^0-9a-zа-яё]+", "", normalize_group_name(value))


def load_group_config() -> dict[str, Any]:
    path = os.getenv("SOHOLMS_GROUP_CONFIG", DEFAULT_CONFIG_PATH)
    if not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as file:
        return json.load(file)


def telegram_chat_ids_from_item(item: dict[str, Any]) -> list[str]:
    raw_chat_ids = item.get("chatIds")
    if raw_chat_ids is None:
        raw_chat_ids = item.get("chat_ids")
    if raw_chat_ids is None:
        raw_chat_ids = [item.get("chatId") or item.get("chat_id")]
    if not isinstance(raw_chat_ids, list):
        raw_chat_ids = [raw_chat_ids]

    chat_ids: list[str] = []
    seen: set[str] = set()
    for raw_chat_id in raw_chat_ids:
        chat_id = normalize_text(raw_chat_id)
        if not chat_id or chat_id in seen:
            continue
        seen.add(chat_id)
        chat_ids.append(chat_id)
    return chat_ids


def telegram_parents_from_item(item: dict[str, Any]) -> list[str]:
    raw_parents = item.get("parents")
    if raw_parents is None:
        raw_parents = item.get("parent")
    if raw_parents is None:
        raw_parents = item.get("parentName")
    if raw_parents is None:
        raw_parents = item.get("parent_name")
    if not isinstance(raw_parents, list):
        raw_parents = [raw_parents]

    parents: list[str] = []
    seen: set[str] = set()
    for raw_parent in raw_parents:
        parent = normalize_text(raw_parent)
        if not parent or parent in seen:
            continue
        seen.add(parent)
        parents.append(parent)
    return parents


def load_split_env_json(prefix: str) -> str:
    direct_value = os.getenv(prefix, "").strip()
    if direct_value:
        return direct_value

    parts: list[str] = []
    for index in range(1, 100):
        value = os.getenv(f"{prefix}_PART_{index}", "")
        if not value:
            break
        parts.append(value)
    return "".join(parts).strip()


def load_telegram_config_items() -> list[dict[str, Any]]:
    path = os.getenv("TELEGRAM_CHAT_CONFIG", DEFAULT_TELEGRAM_CHAT_CONFIG_PATH)
    raw_json = load_split_env_json("TELEGRAM_CHATS_JSON")
    if raw_json:
        raw = json.loads(raw_json)
    elif os.path.exists(path):
        with open(path, "r", encoding="utf-8") as file:
            raw = json.load(file)
    else:
        return {}

    if isinstance(raw, dict) and isinstance(raw.get("students"), list):
        items = raw["students"]
    elif isinstance(raw, list):
        items = raw
    elif isinstance(raw, dict):
        items = [{"name": name, "chatId": chat_id} for name, chat_id in raw.items()]
    else:
        raise BackendError("Unexpected TELEGRAM_CHAT_CONFIG format", HTTPStatus.INTERNAL_SERVER_ERROR)
    return [item for item in items if isinstance(item, dict) and item.get("enabled") is not False]


def load_telegram_chat_targets() -> dict[str, dict[str, Any]]:
    targets: dict[str, dict[str, Any]] = {}
    for item in load_telegram_config_items():
        name = normalize_text(item.get("name"))
        chat_ids = telegram_chat_ids_from_item(item)
        if not name or not chat_ids:
            continue
        key = normalize_group_name(name)
        target = targets.setdefault(key, {"name": name, "chatIds": [], "parents": []})
        for parent in telegram_parents_from_item(item):
            if parent not in target["parents"]:
                target["parents"].append(parent)
        for chat_id in chat_ids:
            if chat_id not in target["chatIds"]:
                target["chatIds"].append(chat_id)
    return targets


def load_telegram_chats() -> dict[str, list[str]]:
    chats: dict[str, list[str]] = {}
    for key, target in load_telegram_chat_targets().items():
        chats[key] = list(target.get("chatIds") or [])
    return chats


def children_by_parent(groups: list[dict[str, Any]]) -> dict[int, list[int]]:
    children: dict[int, list[int]] = defaultdict(list)
    for group in groups:
        group_id = group.get("id")
        if not isinstance(group_id, int):
            continue
        for parent_id in group.get("parentGroupIds") or []:
            if isinstance(parent_id, int):
                children[parent_id].append(group_id)
    return children


def descendant_ids(root_ids: set[int], groups: list[dict[str, Any]]) -> set[int]:
    children = children_by_parent(groups)
    result: set[int] = set()
    stack = list(root_ids)
    while stack:
        group_id = stack.pop()
        for child_id in children.get(group_id, []):
            if child_id in result:
                continue
            result.add(child_id)
            stack.append(child_id)
    return result


def similar_group_names(name: str, groups: list[dict[str, Any]], limit: int = 6) -> list[str]:
    target = compact_group_name(name)
    if not target:
        return []

    scored: list[tuple[float, str]] = []
    seen: set[str] = set()
    for group in groups:
        group_name = normalize_text(group.get("name"))
        if not group_name or group_name in seen:
            continue
        seen.add(group_name)
        compact = compact_group_name(group_name)
        score = SequenceMatcher(None, target, compact).ratio()
        if target in compact or compact in target:
            score += 0.35
        scored.append((score, group_name))

    return [name for score, name in sorted(scored, reverse=True)[:limit] if score >= 0.45]


def resolve_config_group_ids(groups: list[dict[str, Any]], config: dict[str, Any]) -> tuple[set[int], list[str], dict[str, list[str]]]:
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    existing_ids = {group.get("id") for group in groups if isinstance(group.get("id"), int)}
    for group in groups:
        by_name[normalize_group_name(group.get("name"))].append(group)

    selected: set[int] = set()
    missing: list[str] = []
    candidates: dict[str, list[str]] = {}

    for raw_name in config.get("groupNames") or []:
        matches = by_name.get(normalize_group_name(raw_name), [])
        if not matches:
            missing_name = str(raw_name)
            missing.append(missing_name)
            candidates[missing_name] = similar_group_names(missing_name, groups)
            continue
        selected.update(group["id"] for group in matches if isinstance(group.get("id"), int))

    descendant_roots: set[int] = set()
    for raw_name in config.get("includeDescendantsOfNames") or []:
        matches = by_name.get(normalize_group_name(raw_name), [])
        if not matches:
            missing_name = str(raw_name)
            missing.append(missing_name)
            candidates[missing_name] = similar_group_names(missing_name, groups)
            continue
        descendant_roots.update(group["id"] for group in matches if isinstance(group.get("id"), int))
    selected.update(descendant_ids(descendant_roots, groups))

    for group_id in config.get("groupIds") or []:
        try:
            numeric_id = int(group_id)
        except (TypeError, ValueError):
            missing_name = str(group_id)
            missing.append(missing_name)
            candidates[missing_name] = []
            continue
        if numeric_id not in existing_ids:
            missing_name = f"groupId:{numeric_id}"
            missing.append(missing_name)
            candidates[missing_name] = []
            continue
        selected.add(numeric_id)

    return selected, missing, candidates


def teacher_name(teacher: dict[str, Any]) -> str:
    parts = [
        normalize_text(teacher.get("lastName")),
        normalize_text(teacher.get("firstName")),
        normalize_text(teacher.get("middleName")),
    ]
    return " ".join([part for part in parts if part])


def group_teacher_label(group: dict[str, Any]) -> str:
    names: list[str] = []
    seen: set[str] = set()
    for teacher in group.get("teachers") or []:
        if teacher.get("isDisabled"):
            continue
        name = teacher_name(teacher)
        if name and name not in seen:
            seen.add(name)
            names.append(name)
    return ", ".join(names) or "Без преподавателя"


def infer_subject(value: str, parent_subject: str = "") -> str:
    own_text = value.lower()
    for subject, aliases in SUBJECT_ALIASES:
        if any(alias in own_text for alias in aliases):
            return subject
    if parent_subject:
        parent_text = parent_subject.lower()
        for subject, aliases in SUBJECT_ALIASES:
            if any(alias in parent_text for alias in aliases):
                return subject
    return "без предмета"


def infer_level(value: str) -> str:
    text = value.upper()
    if "ОГЭ" in text:
        return "ОГЭ"
    if "ЕГЭ" in text:
        return "ЕГЭ"
    if re.search(r"(^|[^0-9])9([^0-9]|$)", text):
        return "ОГЭ"
    if re.search(r"(^|[^0-9])11([^0-9]|$)", text):
        return "ЕГЭ"
    return ""


def discipline_matches_group(discipline: Any, group: GroupInfo, xlsx_group: str = "") -> bool:
    if xlsx_group and group.subject != "без предмета":
        xlsx_group_subject = infer_subject(normalize_text(xlsx_group))
        if xlsx_group_subject != "без предмета" and xlsx_group_subject != group.subject:
            return False

    text = normalize_text(discipline)
    lowered = text.casefold()
    if not text:
        return True
    if "основн" in lowered:
        return False

    discipline_subject = infer_subject(text)
    if group.subject != "без предмета" and discipline_subject != "без предмета" and discipline_subject != group.subject:
        return False

    expected_level = infer_level(group.name)
    discipline_level = infer_level(text)
    if expected_level and discipline_level and discipline_level != expected_level:
        return False

    return True


def build_group_index(groups: list[dict[str, Any]]) -> dict[int, dict[str, Any]]:
    result = {}
    for group in groups:
        group_id = group.get("id")
        if isinstance(group_id, int):
            result[group_id] = group
    return result


def parent_subject_name(group: dict[str, Any], by_id: dict[int, dict[str, Any]]) -> str:
    for parent_id in group.get("parentGroupIds") or []:
        parent = by_id.get(parent_id)
        if parent:
            subject = infer_subject(parent.get("name", ""))
            if subject != "без предмета":
                return subject
    return ""


def selected_groups(
    group_ids: set[int] | None = None,
    subjects: set[str] | None = None,
    include_virtual: bool = False,
    origins: set[str] | None = None,
) -> list[GroupInfo]:
    groups = fetch_group_tree()
    by_id = build_group_index(groups)
    result: list[GroupInfo] = []

    for group in groups:
        group_id = group.get("id")
        if not isinstance(group_id, int):
            continue
        if group_ids is not None and group_id not in group_ids:
            continue
        if not include_virtual and group.get("purpose") != "PhysicalLearning":
            continue
        if origins and normalize_text(group.get("origin")) not in origins:
            continue
        if not group.get("studentIds"):
            continue

        teacher = group_teacher_label(group)
        if teacher == "Без преподавателя" and group_ids is None:
            continue

        parent_subject = parent_subject_name(group, by_id)
        subject = infer_subject(group.get("name", ""), parent_subject)
        if subjects and subject not in subjects:
            continue

        result.append(
            GroupInfo(
                id=group_id,
                name=normalize_text(group.get("name")),
                subject=subject,
                teacher=teacher,
                parent_group_ids=tuple(group.get("parentGroupIds") or []),
                student_count=len(group.get("studentIds") or []),
            )
        )

    return sorted(result, key=lambda item: (item.subject, item.name, item.id))


def searchable_group_row(group: dict[str, Any], by_id: dict[int, dict[str, Any]]) -> dict[str, Any]:
    parent_subject = parent_subject_name(group, by_id)
    return {
        "id": group.get("id"),
        "name": normalize_text(group.get("name")),
        "subject": infer_subject(group.get("name", ""), parent_subject),
        "teacher": group_teacher_label(group),
        "purpose": group.get("purpose"),
        "origin": group.get("origin"),
        "parent_group_ids": group.get("parentGroupIds") or [],
        "student_count": len(group.get("studentIds") or []),
    }


def search_groups(search: str) -> list[dict[str, Any]]:
    groups = fetch_group_tree()
    by_id = build_group_index(groups)
    query = normalize_group_name(search)
    compact_query = compact_group_name(search)
    result: list[dict[str, Any]] = []

    for group in groups:
        name = group.get("name")
        if not normalize_text(name):
            continue
        normalized_name = normalize_group_name(name)
        compact_name = compact_group_name(name)
        if query in normalized_name or compact_query in compact_name:
            result.append(searchable_group_row(group, by_id))

    return sorted(result, key=lambda item: (item["subject"], item["name"], item["id"] or 0))


def fetch_attendance_xlsx(group_id: int, period_from: str, period_to: str) -> bytes:
    params = {
        "learningGroupId": group_id,
        "periodFrom": period_from,
        "periodTo": period_to,
        "isInteractiveLessons": "true",
        "isAcademicLessons": "true",
        "isStudentsOutOfPeriod": "false",
        "withHomeworks": "true",
        "academicDisciplineIds": "",
        "masterClientIds": "",
    }
    url = f"{API_BASE}{ATTENDANCE_PATH}?{urlencode(params)}"
    last_error: Exception | None = None
    for attempt in range(3):
        try:
            response = requests.get(
                url,
                headers=soholms_headers(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*",
                    kind="excel",
                ),
                timeout=90,
            )
            if response.status_code >= 400:
                raise BackendError(
                    f"Soholms XLSX error {response.status_code} for group {group_id}: {response.text[:500]}",
                    response.status_code,
                )
            return response.content
        except BackendError:
            raise
        except requests.RequestException as error:
            last_error = error
            if attempt < 2:
                time.sleep(1.5 * (attempt + 1))
    raise BackendError(f"Soholms XLSX request failed for group {group_id}: {last_error}", HTTPStatus.BAD_GATEWAY)


def score_value(*values: Any) -> float | None:
    for value in values:
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str) and value.strip():
            match = re.match(r"^\s*(\d+(?:[,.]\d+)?)\s*/\s*(\d+(?:[,.]\d+)?)\s*$", value)
            if match:
                got = float(match.group(1).replace(",", "."))
                total = float(match.group(2).replace(",", "."))
                return 0.0 if total == 0 else got / total * 100
            try:
                return float(value.replace(",", "."))
            except ValueError:
                continue
    return None


def iso_date(value: Any, shift_days: int = 0) -> str:
    if isinstance(value, datetime):
        return (value.date() + timedelta(days=shift_days)).isoformat()
    if isinstance(value, date):
        return (value + timedelta(days=shift_days)).isoformat()
    return normalize_text(value)


def shift_iso_date(value: Any, shift_days: int) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    try:
        return (datetime.fromisoformat(text[:10]).date() + timedelta(days=shift_days)).isoformat()
    except ValueError:
        return text


def attendance_request_period(period_from: str, period_to: str) -> tuple[str, str]:
    if DEADLINE_SHIFT_DAYS == 0:
        return period_from, period_to
    return (
        shift_iso_date(period_from, -DEADLINE_SHIFT_DAYS),
        shift_iso_date(period_to, -DEADLINE_SHIFT_DAYS),
    )


def date_label(value: Any, shift_days: int = 0) -> str:
    if isinstance(value, datetime):
        d = value.date() + timedelta(days=shift_days)
    elif isinstance(value, date):
        d = value + timedelta(days=shift_days)
    else:
        return normalize_text(value)
    return f"{d.day:02d}.{MONTHS_RU.get(d.month, str(d.month))}"


def parse_lesson_number(value: Any) -> int:
    match = re.match(r"^\s*(\d+)\.", normalize_text(value))
    return int(match.group(1)) if match else 0


def is_day_lesson(value: Any) -> bool:
    return "день" in normalize_text(value).lower()


def lesson_day_key(lesson: Any, lesson_date: Any) -> str:
    day_order = parse_lesson_number(lesson)
    if day_order:
        return f"day:{day_order}"
    date_key = iso_date(lesson_date)
    if date_key:
        return f"date:{date_key}"
    return f"lesson:{normalize_text(lesson)}"


def header_index(headers: list[Any], *names: str, default: int | None = None) -> int | None:
    normalized_headers = [normalize_text(header).casefold() for header in headers]
    for name in names:
        normalized_name = normalize_text(name).casefold()
        for index, header in enumerate(normalized_headers):
            if header == normalized_name:
                return index
    return default


def row_value(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def late_days(lesson_date: Any, submitted_at: Any) -> int:
    if not isinstance(lesson_date, (date, datetime)) or not isinstance(submitted_at, (date, datetime)):
        return 0

    lesson_day = lesson_date.date() if isinstance(lesson_date, datetime) else lesson_date
    submitted_day = submitted_at.date() if isinstance(submitted_at, datetime) else submitted_at
    due_day = lesson_day + timedelta(days=DEADLINE_SHIFT_DAYS)
    return max(0, (submitted_day - due_day).days)


def late_penalty(lesson_date: Any, submitted_at: Any) -> int:
    return 1 if late_days(lesson_date, submitted_at) > 0 else 0


def parse_attendance_xlsx(
    content: bytes,
    group: GroupInfo,
    first_attempt_index: dict[tuple[str, str], tuple[datetime, date]] | None = None,
    first_attempt_stats: dict[str, int] | None = None,
) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=3, max_row=3))]
    columns = {
        "student_id": header_index(headers, "ID ученика", default=0),
        "name": header_index(headers, "Ученик", default=1),
        "group": header_index(headers, "Учебная группа", default=2),
        "discipline": header_index(headers, "Дисциплина", default=3),
        "lesson": header_index(headers, "Урок", default=4),
        "lesson_date": header_index(headers, "Дата урока", default=5),
        "lesson_score": header_index(headers, "Оценка за урок", default=8),
        "homework_score": header_index(headers, "Оценка за ДЗ", default=9),
        "checkpoint_score": header_index(headers, "Оценка за СР", default=10),
        "control_score": header_index(headers, "Оценка за КР", default=11),
        "assignment_status": header_index(headers, "Статус сдачи", default=17),
        "assignment_score": header_index(headers, "Оценка", default=18),
        "submitted_at": header_index(headers, "Дата сдачи", default=19),
    }
    students: dict[str, dict[str, Any]] = {}
    group_day_keys: list[str] = []
    group_day_key_set: set[str] = set()
    day_key_orders: dict[str, int] = {}
    group_day_metadata: dict[str, dict[str, Any]] = {}
    current_day: dict[str, Any] | None = None

    def register_group_day(lesson: Any, lesson_date: Any) -> tuple[str, int]:
        raw_day_order = parse_lesson_number(lesson)
        day_key = lesson_day_key(lesson, lesson_date)
        if day_key not in group_day_key_set:
            group_day_key_set.add(day_key)
            group_day_keys.append(day_key)
            day_key_orders[day_key] = raw_day_order or len(group_day_keys)
            group_day_metadata[day_key] = {
                "dateKey": iso_date(lesson_date, shift_days=DEADLINE_SHIFT_DAYS),
                "dateLabel": date_label(lesson_date, shift_days=DEADLINE_SHIFT_DAYS),
                "dateOrder": day_key_orders[day_key],
                "lesson": normalize_text(lesson),
            }
        return day_key, day_key_orders[day_key]

    def has_assignment_submission(assignment_status: Any, assignment_score: Any, submitted_at: Any) -> bool:
        return (
            bool(normalize_text(assignment_status))
            or assignment_score not in (None, "")
            or isinstance(submitted_at, (date, datetime))
        )

    def student_key(value: Any) -> str:
        return str(int(value)) if isinstance(value, (int, float)) else normalize_text(value)

    def first_attempt_for_day(day: dict[str, Any]) -> tuple[datetime, date] | None:
        if not first_attempt_index:
            return None
        lesson_date = day.get("lessonDate")
        if not isinstance(lesson_date, (date, datetime)):
            return None
        if first_attempt_stats is not None:
            first_attempt_stats["lookups"] = first_attempt_stats.get("lookups", 0) + 1

        item = day["item"]
        subject = normalize_text(item.get("subject")).casefold()
        level = normalize_text(item.get("level")).casefold()
        scope = f"{subject}|{level}"

        date_keys = [iso_date(lesson_date, shift_days=DEADLINE_SHIFT_DAYS)]
        student_keys = [
            f"id:{student_key(day.get('studentId'))}|{scope}",
            f"name:{normalize_person_key(item.get('name'))}|{scope}",
        ]
        for date_key in date_keys:
            for key in student_keys:
                entry = first_attempt_index.get((key, date_key))
                if entry:
                    if first_attempt_stats is not None:
                        first_attempt_stats["matched"] = first_attempt_stats.get("matched", 0) + 1
                    return entry
        return None

    def record_submission_penalty(day: dict[str, Any], submitted_at: Any, assignment_score: Any = None) -> None:
        first_attempt_entry = first_attempt_for_day(day)
        if first_attempt_entry is not None:
            first_attempt_at, deadline_at = first_attempt_entry
            lesson_late_days = 0 if first_attempt_at.date() <= deadline_at else late_penalty(day.get("lessonDate"), first_attempt_at)
        else:
            first_attempt_at = None
            lesson_late_days = 0  # нет данных о первой попытке → не штрафуем

        if assignment_score is not None:
            parsed = score_value(assignment_score)
            if parsed is not None:
                day["dailyScore"]["score"] = parsed

        late_by_lesson = day["item"].setdefault("_lateDaysByLesson", {})
        lesson_key = day["dayOrder"]
        previous = late_by_lesson.get(lesson_key)
        if previous is None or lesson_late_days < previous:
            late_by_lesson[lesson_key] = lesson_late_days
            day["dailyScore"]["lateDays"] = lesson_late_days
            if first_attempt_at and first_attempt_stats is not None:
                first_attempt_stats["applied"] = first_attempt_stats.get("applied", 0) + 1
            effective = first_attempt_at or submitted_at
            if isinstance(effective, datetime):
                day["dailyScore"]["submittedAt"] = effective.isoformat()
                if first_attempt_at:
                    day["dailyScore"]["firstAttemptAt"] = first_attempt_at.isoformat()

    all_students_metadata: dict[str, dict[str, Any]] = {}

    for row in worksheet.iter_rows(min_row=4, values_only=True):
        student_id = row_value(row, columns["student_id"])
        name = row_value(row, columns["name"])
        xlsx_group = row_value(row, columns["group"])
        discipline = row_value(row, columns["discipline"])
        lesson = row_value(row, columns["lesson"])
        lesson_date = row_value(row, columns["lesson_date"])
        lesson_score = row_value(row, columns["lesson_score"])
        homework_score = row_value(row, columns["homework_score"])
        checkpoint_score = row_value(row, columns["checkpoint_score"])
        control_score = row_value(row, columns["control_score"])
        assignment_status = row_value(row, columns["assignment_status"])
        assignment_score = row_value(row, columns["assignment_score"])
        submitted_at = row_value(row, columns["submitted_at"])
        score = score_value(checkpoint_score, control_score, homework_score, lesson_score)
        is_potential_day = is_day_lesson(lesson) or isinstance(lesson_date, (date, datetime))
        is_scored_day = score is not None and is_potential_day

        if not name:
            if current_day and has_assignment_submission(assignment_status, assignment_score, submitted_at):
                record_submission_penalty(current_day, submitted_at, assignment_score)
            continue

        _new_key = str(int(student_id)) if isinstance(student_id, (int, float)) else normalize_text(name)
        if (current_day
                and current_day.get("studentKey") == _new_key
                and not is_day_lesson(lesson)
                and has_assignment_submission(assignment_status, assignment_score, submitted_at)):
            record_submission_penalty(current_day, submitted_at, assignment_score)
            continue

        if not discipline_matches_group(discipline, group, xlsx_group or ""):
            current_day = None
            continue

        row_student_key = str(int(student_id)) if isinstance(student_id, (int, float)) else normalize_text(name)
        all_students_metadata[row_student_key] = {
            "subject": group.subject if group.subject != "без предмета" else infer_subject(discipline or xlsx_group or group.name),
            "level": infer_level(discipline) or infer_level(group.name),
            "group": normalize_text(xlsx_group) or group.name,
            "name": normalize_text(name),
            "teacher": group.teacher,
        }

        day_key = ""
        day_order = 0
        if is_potential_day:
            day_key, day_order = register_group_day(lesson, lesson_date)

        if not is_scored_day:
            current_day = None
            continue

        item = students.setdefault(
            row_student_key,
            {
                **all_students_metadata[row_student_key],
                "dailyScores": [],
                "_lateDaysByLesson": {},
            },
        )
        daily_score = {
            "dateKey": iso_date(lesson_date, shift_days=DEADLINE_SHIFT_DAYS),
            "dateLabel": date_label(lesson_date, shift_days=DEADLINE_SHIFT_DAYS),
            "dateOrder": day_order,
            "dayKey": day_key,
            "score": score,
            "completed": True,
            "lateDays": 0,
        }
        item["dailyScores"].append(daily_score)
        current_day = {
            "item": item,
            "studentId": student_id,
            "studentKey": row_student_key,
            "dayOrder": day_order,
            "lessonDate": lesson_date,
            "dailyScore": daily_score,
        }
        if has_assignment_submission(assignment_status, assignment_score, submitted_at):
            record_submission_penalty(current_day, submitted_at, assignment_score)

    for row_student_key, metadata in all_students_metadata.items():
        if row_student_key not in students:
            students[row_student_key] = {
                **metadata,
                "dailyScores": [],
                "_lateDaysByLesson": {},
            }

    rows: list[dict[str, Any]] = []
    group_days_total = len(group_day_keys)
    for item in students.values():
        daily_scores = item["dailyScores"]
        existing_day_keys = {
            str(day.get("dayKey") or day.get("dateOrder") or "")
            for day in daily_scores
            if day.get("dayKey") or day.get("dateOrder")
        }
        for day_key in group_day_keys:
            if day_key in existing_day_keys:
                continue
            metadata = group_day_metadata.get(day_key, {})
            daily_scores.append(
                {
                    "dateKey": metadata.get("dateKey", ""),
                    "dateLabel": metadata.get("dateLabel", ""),
                    "dateOrder": metadata.get("dateOrder") or day_key_orders.get(day_key, 0),
                    "dayKey": day_key,
                    "score": 0,
                    "completed": False,
                    "lateDays": 0,
                }
            )
        daily_scores.sort(key=lambda day: (int(day.get("dateOrder") or 0), normalize_text(day.get("dateKey"))))
        scores_by_day: dict[str, float] = {}
        for day in daily_scores:
            if not day.get("completed", True):
                continue
            day_key = str(day.get("dayKey") or day.get("dateOrder") or "")
            if not day_key:
                continue
            scores_by_day[day_key] = float(day["score"])
        scores = list(scores_by_day.values())
        days_done = len(scores_by_day)
        days_total = group_days_total or days_done
        quality = sum(scores) / days_done if days_done else 0.0
        coefficient = days_done / days_total if days_total else 0.0
        base_score = quality * coefficient
        penalty = float(sum(item.get("_lateDaysByLesson", {}).values()))
        final_score = max(0.0, base_score - penalty)
        item.pop("_lateDaysByLesson", None)
        for day in daily_scores:
            day.pop("dayKey", None)

        item.update(
            {
                "daysDone": days_done,
                "daysTotal": days_total,
                "coefficient": coefficient,
                "quality": quality,
                "baseScore": base_score,
                "penalty": penalty,
                "finalScore": final_score,
                "score": final_score,
            }
        )
        rows.append(item)

    return rows


def inspect_attendance_xlsx(content: bytes, sample_limit: int = 12) -> dict[str, Any]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    worksheet = workbook.active
    headers = [normalize_text(cell.value) for cell in next(worksheet.iter_rows(min_row=3, max_row=3))]
    columns = {
        "name": header_index(headers, "Ученик", default=1),
        "lesson": header_index(headers, "Урок", default=4),
        "lesson_date": header_index(headers, "Дата урока", default=5),
        "assignment_status": header_index(headers, "Статус сдачи", default=17),
        "assignment_score": header_index(headers, "Оценка", default=18),
        "submitted_at": header_index(headers, "Дата сдачи", default=19),
    }
    total_rows = 0
    day_rows = 0
    assignment_rows = 0
    submitted_rows = 0
    named_assignment_rows = 0
    unnamed_assignment_rows = 0
    samples: list[dict[str, Any]] = []

    for row in worksheet.iter_rows(min_row=4, values_only=True):
        total_rows += 1
        name = row_value(row, columns["name"])
        lesson = row_value(row, columns["lesson"])
        status = row_value(row, columns["assignment_status"])
        assignment_score = row_value(row, columns["assignment_score"])
        submitted_at = row_value(row, columns["submitted_at"])
        has_assignment = bool(normalize_text(status)) or assignment_score not in (None, "")

        if is_day_lesson(lesson):
            day_rows += 1
        if has_assignment:
            assignment_rows += 1
            if name:
                named_assignment_rows += 1
            else:
                unnamed_assignment_rows += 1
        if submitted_at:
            submitted_rows += 1
        if has_assignment and len(samples) < sample_limit:
            samples.append({
                "name": normalize_text(name),
                "lesson": normalize_text(lesson),
                "lessonDate": iso_date(row_value(row, columns["lesson_date"])),
                "status": normalize_text(status),
                "assignmentScore": normalize_text(assignment_score),
                "submittedAt": iso_date(submitted_at),
            })

    return {
        "headers": headers,
        "columns": columns,
        "stats": {
            "totalRows": total_rows,
            "dayRows": day_rows,
            "assignmentRows": assignment_rows,
            "submittedRows": submitted_rows,
            "namedAssignmentRows": named_assignment_rows,
            "unnamedAssignmentRows": unnamed_assignment_rows,
        },
        "samples": samples,
    }


_MONTH_RU = {
    "01": "Январь", "02": "Февраль", "03": "Март",
    "04": "Апрель", "05": "Май", "06": "Июнь",
    "07": "Июль", "08": "Август", "09": "Сентябрь",
    "10": "Октябрь", "11": "Ноябрь", "12": "Декабрь",
}


def _month_label(mk: str) -> str:
    if len(mk) >= 7:
        return f"{_MONTH_RU.get(mk[5:7], mk[5:7])} {mk[:4]}"
    return mk


def parse_xlsx_raw_rows(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=3, max_row=3))]
    columns = {
        "name": header_index(headers, "Ученик", default=1),
        "group": header_index(headers, "Учебная группа", default=2),
        "discipline": header_index(headers, "Дисциплина", default=3),
        "lesson_date": header_index(headers, "Дата урока", default=5),
        "lesson_score": header_index(headers, "Оценка за урок", default=8),
        "homework_score": header_index(headers, "Оценка за ДЗ", default=9),
        "checkpoint_score": header_index(headers, "Оценка за СР", default=10),
        "control_score": header_index(headers, "Оценка за КР", default=11),
    }
    rows: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=4, values_only=True):
        name = row_value(row, columns["name"])
        if not name:
            continue
        discipline = normalize_text(row_value(row, columns["discipline"]))
        if discipline and "основн" not in discipline.casefold():
            continue
        date_key = iso_date(row_value(row, columns["lesson_date"]))
        if len(date_key) < 7:
            continue
        rows.append({
            "name": normalize_text(name),
            "group": normalize_text(row_value(row, columns["group"])),
            "month_key": date_key[:7],
            "hw": score_value(row_value(row, columns["homework_score"])),
            "sr": score_value(row_value(row, columns["checkpoint_score"])),
            "kr": score_value(row_value(row, columns["control_score"])),
            "lesson": score_value(row_value(row, columns["lesson_score"])),
        })
    return rows


def fetch_year_raw_rows(period_from: str, period_to: str) -> list[dict[str, Any]]:
    cache_key = f"year_raw_rows:{period_from}:{period_to}"

    def load():
        groups = selected_groups()
        all_rows: list[dict[str, Any]] = []

        def fetch_group(group: GroupInfo):
            content = fetch_attendance_xlsx(group.id, period_from, period_to)
            rows = parse_xlsx_raw_rows(content)
            level = infer_level(group.name) or ""
            for r in rows:
                r["_subject"] = group.subject
                r["_level"] = level
                r["_teacher"] = group.teacher
            return rows

        with ThreadPoolExecutor(max_workers=max(1, DEFAULT_CONCURRENCY)) as executor:
            futures = {executor.submit(fetch_group, group): group for group in groups}
            for future in as_completed(futures):
                try:
                    all_rows.extend(future.result())
                except Exception:
                    pass
        return all_rows

    return cached(cache_key, 3600, load)


def aggregate_student_year(rows: list[dict[str, Any]], student_name: str) -> dict[str, Any] | None:
    name_key = normalize_person_key(student_name)
    student_rows = [r for r in rows if normalize_person_key(r["name"]).startswith(name_key)]
    if not student_rows:
        return None

    months_data: dict[str, dict[str, Any]] = {}
    for r in student_rows:
        mk = r["month_key"]
        if mk not in months_data:
            months_data[mk] = {"hw_done": 0, "hw_total": 0, "test_scores": [], "lessons": 0, "present": 0}
        m = months_data[mk]
        m["lessons"] += 1
        if r["lesson"] is not None or r["sr"] is not None or r["kr"] is not None or r["hw"] is not None:
            m["present"] += 1
        if r["hw"] is not None:
            m["hw_total"] += 1
            if r["hw"] > 0:
                m["hw_done"] += 1
        if r["kr"] is not None:
            m["test_scores"].append(r["kr"])
        if r["sr"] is not None:
            m["test_scores"].append(r["sr"])

    months = []
    for mk in sorted(months_data):
        m = months_data[mk]
        test_avg = round(sum(m["test_scores"]) / len(m["test_scores"]), 1) if m["test_scores"] else None
        attendance = round(m["present"] / m["lessons"] * 100) if m["lessons"] else 0
        months.append({
            "month": _month_label(mk),
            "hwDone": m["hw_done"],
            "hwTotal": m["hw_total"],
            "testAvg": test_avg,
            "attendance": attendance,
        })

    total_hw_done = sum(m["hwDone"] for m in months)
    total_hw_total = sum(m["hwTotal"] for m in months)
    test_avgs = [m["testAvg"] for m in months if m["testAvg"] is not None]
    overall_test_avg = round(sum(test_avgs) / len(test_avgs), 1) if test_avgs else None
    attendance_vals = [m["attendance"] for m in months]
    avg_attendance = round(sum(attendance_vals) / len(attendance_vals)) if attendance_vals else 0

    return {
        "months": months,
        "summary": {
            "hwDone": total_hw_done,
            "hwTotal": total_hw_total,
            "testAvg": overall_test_avg,
            "attendancePct": avg_attendance,
        },
    }


def aggregate_all_students(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    students: dict[tuple[str, str], dict[str, Any]] = {}

    for r in rows:
        name = r.get("name", "")
        group = r.get("group", "")
        if not name:
            continue
        key = (normalize_person_key(name), group)
        if key not in students:
            students[key] = {
                "name": name,
                "group": group,
                "subject": r.get("_subject", ""),
                "level": r.get("_level", ""),
                "teacher": r.get("_teacher", ""),
                "hw_scores": [],
                "kr_scores": [],
                "lessons": 0,
                "present": 0,
            }
        s = students[key]
        s["lessons"] += 1
        hw = r.get("hw")
        kr = r.get("kr")
        sr = r.get("sr")
        lesson = r.get("lesson")
        if hw is not None or kr is not None or sr is not None or lesson is not None:
            s["present"] += 1
        if hw is not None:
            s["hw_scores"].append(hw)
        kr_val = kr if kr is not None else sr
        if kr_val is not None:
            s["kr_scores"].append(kr_val)

    result = []
    for s in students.values():
        hw_scores = s["hw_scores"]
        kr_scores = s["kr_scores"]
        hw_avg = round(sum(hw_scores) / len(hw_scores), 1) if hw_scores else None
        kr_avg = round(sum(kr_scores) / len(kr_scores), 1) if kr_scores else None
        attendance = round(s["present"] / s["lessons"] * 100) if s["lessons"] else 0
        result.append({
            "name": s["name"],
            "group": s["group"],
            "subject": s["subject"],
            "level": s["level"],
            "teacher": s["teacher"],
            "hwAvg": hw_avg,
            "hwCount": len(hw_scores),
            "krAvg": kr_avg,
            "krCount": len(kr_scores),
            "attendancePct": attendance,
            "lessonsTotal": s["lessons"],
        })

    return sorted(result, key=lambda x: (x.get("group", ""), x.get("name", "")))


# === Электронный журнал: полный парсер + per-student / school агрегаты ===


def parse_journal_full(content: bytes) -> list[dict[str, Any]]:
    """Полный парсер электронного журнала Soholms.

    Возвращает список событий (заголовков с ID ученика). К каждому событию
    приклеены под-строки заданий (sub_rows). Фильтрация: только дисциплина
    «Основной курс по ...».
    """
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=3, max_row=3))]
    cols = {
        "id":         header_index(headers, "ID ученика", default=0),
        "name":       header_index(headers, "Ученик", default=1),
        "group":      header_index(headers, "Учебная группа", default=2),
        "discipline": header_index(headers, "Дисциплина", default=3),
        "lesson":     header_index(headers, "Урок", default=4),
        "date_l":     header_index(headers, "Дата урока", default=5),
        "date_a":     header_index(headers, "Дата доступности урока", default=6),
        "status":     header_index(headers, "Статус", default=7),
        "g_lesson":   header_index(headers, "Оценка за урок", default=8),
        "g_dz":       header_index(headers, "Оценка за ДЗ", default=9),
        "g_sr":       header_index(headers, "Оценка за СР", default=10),
        "g_kr":       header_index(headers, "Оценка за КР", default=11),
        "task_kind":  header_index(headers, "Вид задания", default=15),
        "task_class": header_index(headers, "Тип задания", default=16),
        "sub_status": header_index(headers, "Статус сдачи", default=17),
    }
    events: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for row in worksheet.iter_rows(min_row=4, values_only=True):
        sid = row_value(row, cols["id"])
        if sid:
            discipline = normalize_text(row_value(row, cols["discipline"]))
            if discipline and not discipline.casefold().startswith("основной курс"):
                current = None
                continue
            lesson = normalize_text(row_value(row, cols["lesson"]))
            ev_type = lesson.split()[0] if lesson else ""
            date_raw = row_value(row, cols["date_l"]) or row_value(row, cols["date_a"])
            current = {
                "student": normalize_text(row_value(row, cols["name"])),
                "group":   normalize_text(row_value(row, cols["group"])),
                "discipline": discipline,
                "lesson":  lesson,
                "type":    ev_type,
                "date":    iso_date(date_raw) if date_raw else "",
                "status":  normalize_text(row_value(row, cols["status"])),
                "g_lesson": score_value(row_value(row, cols["g_lesson"])),
                "g_dz":    score_value(row_value(row, cols["g_dz"])),
                "g_sr":    score_value(row_value(row, cols["g_sr"])),
                "g_kr":    score_value(row_value(row, cols["g_kr"])),
                "sub_rows": [],
            }
            events.append(current)
        else:
            if current is None:
                continue
            current["sub_rows"].append({
                "task_kind":  normalize_text(row_value(row, cols["task_kind"])),
                "task_class": normalize_text(row_value(row, cols["task_class"])),
                "status":     normalize_text(row_value(row, cols["sub_status"])),
            })
    return events


def fetch_journal_events(period_from: str, period_to: str) -> list[dict[str, Any]]:
    """Параллельная загрузка журнала по всем группам. Кеш 1 час."""
    cache_key = f"journal_events:{period_from}:{period_to}"

    def load() -> list[dict[str, Any]]:
        groups = selected_groups()
        all_events: list[dict[str, Any]] = []

        def fetch_group(group: GroupInfo) -> list[dict[str, Any]]:
            content = fetch_attendance_xlsx(group.id, period_from, period_to)
            evs = parse_journal_full(content)
            level = infer_level(group.name) or ""
            for e in evs:
                e["_subject"] = group.subject
                e["_teacher"] = group.teacher
                e["_level"] = level
                e["_group_id"] = group.id
            return evs

        with ThreadPoolExecutor(max_workers=max(1, DEFAULT_CONCURRENCY)) as executor:
            futures = {executor.submit(fetch_group, group): group for group in groups}
            for future in as_completed(futures):
                try:
                    all_events.extend(future.result())
                except Exception:
                    pass
        return all_events

    return cached(cache_key, 3600, load)


def _journal_max_per_lesson(events: list[dict[str, Any]]) -> tuple[dict, dict]:
    hw_max: dict[tuple[str, str], float] = {}
    kr_max: dict[tuple[str, str], float] = {}
    for e in events:
        key = (e.get("group", ""), e.get("lesson", ""))
        if e.get("type") == "Домашка" and isinstance(e.get("g_dz"), (int, float)):
            v = float(e["g_dz"])
            if v > hw_max.get(key, 0):
                hw_max[key] = v
        if e.get("type") == "Занятие" and isinstance(e.get("g_kr"), (int, float)):
            v = float(e["g_kr"])
            if v > kr_max.get(key, 0):
                kr_max[key] = v
    return hw_max, kr_max


def _journal_extract_no(lesson: str, prefix: str) -> int | None:
    if not lesson:
        return None
    match = re.search(rf"{prefix}\s*№?\s*(\d+)", lesson)
    return int(match.group(1)) if match else None


def _journal_aggregate_one(
    student_events: list[dict[str, Any]],
    hw_max: dict[tuple[str, str], float],
    kr_max: dict[tuple[str, str], float],
) -> dict[str, Any]:
    """Считает посещаемость / ДЗ / КР / интеграл / флаги для одного ученика."""
    # Attendance — события типа "Математика"
    att_events = [e for e in student_events if e["type"] == "Математика"]
    att_sorted = sorted(
        [e for e in att_events if e["status"] in ("Был", "Не был", "Болел")],
        key=lambda x: x.get("date") or "",
    )
    counts = {"Был": 0, "Не был": 0, "Болел": 0}
    for e in att_sorted:
        counts[e["status"]] += 1
    total = sum(counts.values())
    att_pct = round(counts["Был"] / total * 100, 1) if total else None
    max_streak = 0
    streak = 0
    for e in att_sorted:
        if e["status"] == "Не был":
            streak += 1
            if streak > max_streak:
                max_streak = streak
        else:
            streak = 0

    by_month: dict[str, dict[str, int]] = {}
    for e in att_sorted:
        mk = (e.get("date") or "")[:7]
        if not mk:
            continue
        m = by_month.setdefault(mk, {"was": 0, "absent": 0, "sick": 0})
        if e["status"] == "Был":
            m["was"] += 1
        elif e["status"] == "Не был":
            m["absent"] += 1
        elif e["status"] == "Болел":
            m["sick"] += 1
    months_list = []
    for mk in sorted(by_month):
        m = by_month[mk]
        tot = m["was"] + m["absent"] + m["sick"]
        months_list.append({
            "month": _month_label(mk),
            "monthKey": mk,
            "was": m["was"],
            "absent": m["absent"],
            "sick": m["sick"],
            "total": tot,
            "pct": round(m["was"] / tot * 100, 1) if tot else 0,
        })

    # Homework — нормализуем по max в (group, lesson)
    hw_events = [e for e in student_events if e["type"] == "Домашка"]
    hw_norm: list[float] = []
    hw_trend: list[dict[str, Any]] = []
    sub_status_count: dict[str, int] = {}
    for e in hw_events:
        if isinstance(e.get("g_dz"), (int, float)):
            mx = hw_max.get((e["group"], e["lesson"]))
            if mx and mx > 0:
                pct = float(e["g_dz"]) / mx * 100
                hw_norm.append(pct)
                hw_trend.append({
                    "no": _journal_extract_no(e["lesson"], "Домашка"),
                    "lesson": e["lesson"],
                    "grade": round(pct, 1),
                    "raw": float(e["g_dz"]),
                    "max": mx,
                    "date": e.get("date") or "",
                })
        for sr in e["sub_rows"]:
            st = sr.get("status") or ""
            if st:
                sub_status_count[st] = sub_status_count.get(st, 0) + 1
    hw_total_subs = sum(sub_status_count.values())
    hw_done_pct = (
        round(sub_status_count.get("Принято", 0) / hw_total_subs * 100, 1)
        if hw_total_subs else None
    )
    hw_not_opened_pct = (
        round(sub_status_count.get("Не открывали", 0) / hw_total_subs * 100, 1)
        if hw_total_subs else 0
    )
    hw_avg = round(sum(hw_norm) / len(hw_norm), 1) if hw_norm else None
    hw_trend.sort(key=lambda x: (x["no"] is None, x["no"] or 0, x["date"] or ""))

    # КР: g_kr на строках "Занятие"
    kr_records: list[dict[str, Any]] = []
    for e in student_events:
        if e["type"] == "Занятие" and isinstance(e.get("g_kr"), (int, float)):
            mx = kr_max.get((e["group"], e["lesson"]))
            if mx and mx > 0:
                pct = float(e["g_kr"]) / mx * 100
                kr_records.append({
                    "lesson": e["lesson"],
                    "grade": round(pct, 1),
                    "raw": float(e["g_kr"]),
                    "max": mx,
                    "date": e.get("date") or "",
                })
    kr_records.sort(key=lambda x: x["date"] or "")
    kr_avg = (
        round(sum(k["grade"] for k in kr_records) / len(kr_records), 1)
        if kr_records else None
    )

    # Флаги
    flags: list[str] = []
    if max_streak >= 3:
        flags.append(f"Пропусков подряд: {max_streak}")
    if hw_total_subs and hw_not_opened_pct >= 30:
        flags.append(f"Не открыто {round(hw_not_opened_pct)}% заданий ДЗ")
    if hw_avg is not None and hw_avg < 60:
        flags.append(f"Средний ДЗ {round(hw_avg)}% (низкий)")
    if kr_avg is not None and kr_avg < 60:
        flags.append(f"Средний КР {round(kr_avg)}% (низкий)")
    if att_pct is not None and att_pct < 70:
        flags.append(f"Посещаемость {round(att_pct)}% (низкая)")

    components = [v for v in (att_pct, hw_avg, kr_avg) if v is not None]
    integral = round(sum(components) / len(components), 1) if components else None

    return {
        "attendance": {
            "pct": att_pct,
            "was": counts["Был"],
            "absent": counts["Не был"],
            "sick": counts["Болел"],
            "total": total,
            "max_streak": max_streak,
            "by_month": months_list,
        },
        "homework": {
            "avg": hw_avg,
            "count": len(hw_norm),
            "done_pct": hw_done_pct,
            "not_opened_pct": hw_not_opened_pct,
            "sub_status": sub_status_count,
            "trend": hw_trend,
        },
        "kr": {
            "avg": kr_avg,
            "count": len(kr_records),
            "list": kr_records,
        },
        "integral": integral,
        "flags": flags,
    }


def _journal_group_by_student(events: list[dict[str, Any]]) -> dict[tuple[str, str], dict[str, Any]]:
    """Сгруппировать события по (norm_name, group) — один ученик в группе."""
    by_student: dict[tuple[str, str], dict[str, Any]] = {}
    for e in events:
        sn = e.get("student") or ""
        if not sn:
            continue
        key = (normalize_person_key(sn), e.get("group", ""))
        if key not in by_student:
            by_student[key] = {
                "name": sn,
                "group": e.get("group", ""),
                "subject": e.get("_subject", ""),
                "teacher": e.get("_teacher", ""),
                "level": e.get("_level", ""),
                "events": [],
            }
        by_student[key]["events"].append(e)
    return by_student


def aggregate_student_journal(
    events: list[dict[str, Any]], student_name: str, subject: str = ""
) -> dict[str, Any] | None:
    name_key = normalize_person_key(student_name)
    if not name_key:
        return None
    by_student = _journal_group_by_student(events)
    matching_keys = [k for k in by_student if k[0].startswith(name_key)]
    if not matching_keys:
        return None
    # If subject specified — prefer entries matching that subject
    if subject:
        subj_norm = subject.strip().lower()
        subj_filtered = [k for k in matching_keys if by_student[k].get("subject", "").lower() == subj_norm]
        if subj_filtered:
            matching_keys = subj_filtered
    # Among remaining, pick entry with most events (most complete journal data)
    target_key = max(matching_keys, key=lambda k: len(by_student[k]["events"]))

    hw_max, kr_max = _journal_max_per_lesson(events)
    target = by_student[target_key]
    detail = _journal_aggregate_one(target["events"], hw_max, kr_max)
    detail["name"] = target["name"]
    detail["group"] = target["group"]
    detail["subject"] = target["subject"]
    detail["teacher"] = target["teacher"]

    integrals = []
    for key, data in by_student.items():
        agg = _journal_aggregate_one(data["events"], hw_max, kr_max)
        if agg["integral"] is not None:
            integrals.append({
                "name_key": key[0],
                "group": data["group"],
                "integral": agg["integral"],
            })

    same_group = sorted(
        [x for x in integrals if x["group"] == target["group"]],
        key=lambda x: -x["integral"],
    )
    school = sorted(integrals, key=lambda x: -x["integral"])

    def _place(lst, key):
        for i, x in enumerate(lst):
            if x["name_key"] == key:
                return i + 1
        return None

    detail["ranks"] = {
        "in_group": {"place": _place(same_group, target_key[0]), "of": len(same_group)},
        "in_school": {"place": _place(school, target_key[0]), "of": len(school)},
    }
    return detail


def aggregate_school_journal(events: list[dict[str, Any]], mode: str) -> dict[str, Any]:
    hw_max, kr_max = _journal_max_per_lesson(events)
    by_student = _journal_group_by_student(events)

    students_agg: list[dict[str, Any]] = []
    for data in by_student.values():
        agg = _journal_aggregate_one(data["events"], hw_max, kr_max)
        students_agg.append({
            "name": data["name"],
            "group": data["group"],
            "subject": data["subject"],
            "teacher": data["teacher"],
            "level": data["level"],
            "attendancePct": agg["attendance"]["pct"],
            "hwAvg": agg["homework"]["avg"],
            "hwCount": agg["homework"]["count"],
            "krAvg": agg["kr"]["avg"],
            "krCount": agg["kr"]["count"],
            "integral": agg["integral"],
            "flags": agg["flags"],
            "absent": agg["attendance"]["absent"] + agg["attendance"]["sick"],
            "maxStreak": agg["attendance"]["max_streak"],
            "lessonsTotal": agg["attendance"]["total"],
        })

    if mode == "students":
        students_agg.sort(key=lambda x: (x["group"] or "", x["name"] or ""))
        return {"students": students_agg}

    groups_map: dict[str, dict[str, Any]] = {}
    for s in students_agg:
        g = s["group"]
        gm = groups_map.setdefault(g, {
            "group": g, "subject": s["subject"], "teacher": s["teacher"], "level": s["level"],
            "students_count": 0, "att": [], "hw": [], "kr": [], "integral": [], "flags_count": 0,
        })
        gm["students_count"] += 1
        if s["attendancePct"] is not None: gm["att"].append(s["attendancePct"])
        if s["hwAvg"] is not None: gm["hw"].append(s["hwAvg"])
        if s["krAvg"] is not None: gm["kr"].append(s["krAvg"])
        if s["integral"] is not None: gm["integral"].append(s["integral"])
        gm["flags_count"] += len(s["flags"])

    def _avg(lst):
        return round(sum(lst) / len(lst), 1) if lst else None

    groups_list = [{
        "group": gm["group"],
        "subject": gm["subject"],
        "teacher": gm["teacher"],
        "level": gm["level"],
        "studentsCount": gm["students_count"],
        "attendancePct": _avg(gm["att"]),
        "hwAvg": _avg(gm["hw"]),
        "krAvg": _avg(gm["kr"]),
        "integral": _avg(gm["integral"]),
        "flagsCount": gm["flags_count"],
    } for gm in groups_map.values()]
    groups_list.sort(key=lambda x: -(x["integral"] if x["integral"] is not None else -1))

    if mode == "groups":
        return {"groups": groups_list}

    # mode == "school"
    all_att = [s["attendancePct"] for s in students_agg if s["attendancePct"] is not None]
    all_hw = [s["hwAvg"] for s in students_agg if s["hwAvg"] is not None]
    all_kr = [s["krAvg"] for s in students_agg if s["krAvg"] is not None]
    all_int = [s["integral"] for s in students_agg if s["integral"] is not None]
    students_sorted = sorted(
        [s for s in students_agg if s["integral"] is not None],
        key=lambda x: -x["integral"],
    )
    return {
        "studentsCount": len(students_agg),
        "groupsCount": len(groups_map),
        "attendancePct": _avg(all_att),
        "hwAvg": _avg(all_hw),
        "krAvg": _avg(all_kr),
        "integral": _avg(all_int),
        "flagsCount": sum(len(s["flags"]) for s in students_agg),
        "top3Groups": groups_list[:3],
        "top3Students": [
            {"name": s["name"], "group": s["group"], "integral": s["integral"]}
            for s in students_sorted[:3]
        ],
        "bottom3Students": [
            {"name": s["name"], "group": s["group"], "integral": s["integral"]}
            for s in students_sorted[-3:][::-1]
        ] if len(students_sorted) >= 3 else [],
    }


# === конец блока электронного журнала ===


def add_places(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_group: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    by_school: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)

    for row in rows:
        by_group[(row.get("subject", ""), row.get("level", ""), row.get("group", ""))].append(row)
        by_school[(row.get("subject", ""), row.get("level", ""))].append(row)

    def apply_place(items: list[dict[str, Any]], key: str) -> None:
        sorted_items = sorted(items, key=lambda row: float(row.get("finalScore") or 0), reverse=True)
        previous_score = None
        place = 0
        for item in sorted_items:
            score = float(item.get("finalScore") or 0)
            if score != previous_score:
                place += 1
                previous_score = score
            item[key] = place

    for items in by_group.values():
        apply_place(items, "groupPlace")
    for items in by_school.values():
        apply_place(items, "schoolPlace")

    return rows


def student_row_identity(row: dict[str, Any]) -> str:
    parts = [
        row.get("subject", ""),
        row.get("level", ""),
        row.get("group", ""),
        row.get("name", ""),
        row.get("teacher", ""),
    ]
    return "\x1f".join(normalize_text(part).casefold() for part in parts)


def dedupe_rating_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Keep one row per (name, subject, level, group); prefer highest finalScore."""
    seen: dict[tuple[str, str, str, str], int] = {}
    for i, row in enumerate(rows):
        key = (
            normalize_text(row.get("name", "")).casefold(),
            normalize_text(row.get("subject", "")).casefold(),
            normalize_text(row.get("level", "")).casefold(),
            normalize_text(row.get("group", "")).casefold(),
        )
        prev = seen.get(key)
        if prev is None:
            seen[key] = i
        elif float(row.get("finalScore") or 0) > float(rows[prev].get("finalScore") or 0):
            seen[key] = i
    return [rows[i] for i in sorted(seen.values())]


def penalty_override_key(period: dict[str, Any], row: dict[str, Any]) -> str:
    period_from = normalize_text(period.get("from"))
    period_to = normalize_text(period.get("to"))
    return "\x1f".join([period_from, period_to, student_row_identity(row)])


def load_penalty_overrides() -> dict[str, dict[str, Any]]:
    if not os.path.exists(PENALTY_OVERRIDES_PATH):
        return {}
    with open(PENALTY_OVERRIDES_PATH, "r", encoding="utf-8") as file:
        data = json.load(file)
    overrides = data.get("overrides", data) if isinstance(data, dict) else {}
    return overrides if isinstance(overrides, dict) else {}


def write_penalty_overrides(overrides: dict[str, dict[str, Any]]) -> None:
    directory = os.path.dirname(PENALTY_OVERRIDES_PATH)
    if directory:
        os.makedirs(directory, exist_ok=True)
    temp_path = f"{PENALTY_OVERRIDES_PATH}.tmp"
    with open(temp_path, "w", encoding="utf-8") as file:
        json.dump({"overrides": overrides}, file, ensure_ascii=False, indent=2, sort_keys=True)
        file.write("\n")
    os.replace(temp_path, PENALTY_OVERRIDES_PATH)


def parse_penalty_value(value: Any, field_name: str = "penalty") -> float:
    try:
        penalty = float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        raise BackendError(f"{field_name} must be a number", HTTPStatus.BAD_REQUEST)
    if penalty < 0:
        raise BackendError(f"{field_name} must be non-negative", HTTPStatus.BAD_REQUEST)
    if penalty > 1000:
        raise BackendError(f"{field_name} is too large", HTTPStatus.BAD_REQUEST)
    return penalty


def apply_penalty_overrides(payload: dict[str, Any]) -> dict[str, Any]:
    rows = payload.get("rows", [])
    if not isinstance(rows, list):
        return payload

    period = payload.get("period", {})
    if not isinstance(period, dict):
        period = {}

    with _PENALTY_OVERRIDES_LOCK:
        overrides = load_penalty_overrides()

    for row in rows:
        if not isinstance(row, dict):
            continue

        key = penalty_override_key(period, row)
        auto_penalty = float(row.get("penalty") or 0)
        row["overrideKey"] = key
        row["autoPenalty"] = auto_penalty
        row["penaltyOverridden"] = False

        override = overrides.get(key)
        if isinstance(override, dict) and "penalty" in override:
            penalty = parse_penalty_value(override.get("penalty"))
            row["penalty"] = penalty
            row["penaltyOverridden"] = True
            final_score = max(0.0, float(row.get("baseScore") or 0) - penalty)
            row["finalScore"] = final_score
            row["score"] = final_score

    add_places(rows)
    return payload


def save_penalty_override(body: dict[str, Any]) -> dict[str, Any]:
    row = {
        "subject": body.get("subject", ""),
        "level": body.get("level", ""),
        "group": body.get("group", ""),
        "name": body.get("name", ""),
        "teacher": body.get("teacher", ""),
    }
    period = {
        "from": body.get("periodFrom", ""),
        "to": body.get("periodTo", ""),
    }

    if not normalize_text(row["name"]):
        raise BackendError("name is required", HTTPStatus.BAD_REQUEST)
    if not normalize_text(period["from"]) or not normalize_text(period["to"]):
        raise BackendError("periodFrom and periodTo are required", HTTPStatus.BAD_REQUEST)

    penalty = parse_penalty_value(body.get("penalty"))
    auto_penalty = parse_penalty_value(body.get("autoPenalty", penalty), "autoPenalty")
    key = penalty_override_key(period, row)
    now = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"

    with _PENALTY_OVERRIDES_LOCK:
        overrides = load_penalty_overrides()
        if abs(penalty - auto_penalty) < 0.000001:
            overrides.pop(key, None)
            overridden = False
        else:
            overrides[key] = {
                "penalty": penalty,
                "updatedAt": now,
                "row": row,
                "period": period,
            }
            overridden = True
        write_penalty_overrides(overrides)

    return {
        "ok": True,
        "overrideKey": key,
        "penalty": penalty,
        "autoPenalty": auto_penalty,
        "penaltyOverridden": overridden,
    }


def public_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "subject": row.get("subject", ""),
        "level": row.get("level", ""),
        "group": row.get("group", ""),
        "name": row.get("name", ""),
        "teacher": row.get("teacher", "Без преподавателя"),
        "score": row.get("score", 0),
        "finalScore": row.get("finalScore", row.get("score", 0)),
        "groupPlace": row.get("groupPlace", 0),
        "schoolPlace": row.get("schoolPlace", 0),
    }


def strip_for_public(payload: dict[str, Any]) -> dict[str, Any]:
    return {
        **payload,
        "rows": [public_row(row) for row in payload.get("rows", [])],
    }


PUBLIC_RATINGS_QUERY_KEYS = (
    "periodFrom",
    "periodTo",
    "groupIds",
    "subjects",
    "includeVirtual",
    "origins",
    "limit",
)


def normalize_public_ratings_query(query: dict[str, str] | None = None) -> dict[str, str]:
    query = query or {}
    default_from, default_to = current_marathon_period()
    normalized = {
        "periodFrom": default_from,
        "periodTo": default_to,
    }
    normalized.update({
        key: str(query.get(key, "")).strip()
        for key in PUBLIC_RATINGS_QUERY_KEYS
        if str(query.get(key, "")).strip()
    })
    return {key: value for key, value in normalized.items() if value}


def normalize_admin_ratings_query(query: dict[str, str] | None = None) -> dict[str, str]:
    query = query or {}
    default_from, default_to = current_marathon_period()
    normalized = {
        "periodFrom": query.get("periodFrom") or os.getenv("SOHOLMS_PERIOD_FROM") or default_from,
        "periodTo": query.get("periodTo") or os.getenv("SOHOLMS_PERIOD_TO") or default_to,
    }
    normalized.update({
        key: str(query.get(key, "")).strip()
        for key in PUBLIC_RATINGS_QUERY_KEYS
        if str(query.get(key, "")).strip()
    })
    return {key: str(value).strip() for key, value in normalized.items() if str(value).strip()}


def public_ratings_refresh_key(query: dict[str, str] | None = None) -> str:
    return json.dumps(normalize_public_ratings_query(query), ensure_ascii=False, sort_keys=True)


def admin_ratings_refresh_key(query: dict[str, str] | None = None) -> str:
    return json.dumps(normalize_admin_ratings_query(query), ensure_ascii=False, sort_keys=True)


def utc_timestamp() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def snapshot_next_refresh_at(saved_at_epoch: float, refresh_seconds: int) -> str:
    return datetime.utcfromtimestamp(saved_at_epoch + refresh_seconds).replace(microsecond=0).isoformat() + "Z"


def read_public_ratings_snapshot_file() -> dict[str, Any] | None:
    if not os.path.exists(PUBLIC_RATINGS_SNAPSHOT_PATH):
        return None

    with _PUBLIC_RATINGS_SNAPSHOT_LOCK:
        try:
            with open(PUBLIC_RATINGS_SNAPSHOT_PATH, "r", encoding="utf-8") as file:
                snapshot = json.load(file)
        except (OSError, json.JSONDecodeError):
            return None

    return snapshot if isinstance(snapshot, dict) else None


def write_public_ratings_snapshot_file(snapshot: dict[str, Any]) -> None:
    directory = os.path.dirname(PUBLIC_RATINGS_SNAPSHOT_PATH)
    if directory:
        os.makedirs(directory, exist_ok=True)
    temp_path = f"{PUBLIC_RATINGS_SNAPSHOT_PATH}.tmp"

    with _PUBLIC_RATINGS_SNAPSHOT_LOCK:
        with open(temp_path, "w", encoding="utf-8") as file:
            json.dump(snapshot, file, ensure_ascii=False, separators=(",", ":"))
        os.replace(temp_path, PUBLIC_RATINGS_SNAPSHOT_PATH)


def add_public_snapshot_meta(payload: dict[str, Any], snapshot: dict[str, Any]) -> dict[str, Any]:
    saved_at_epoch = float(snapshot.get("savedAtEpoch") or 0)
    return {
        **copy.deepcopy(strip_for_public(payload)),
        "snapshot": {
            "updatedAt": snapshot.get("savedAt", ""),
            "nextRefreshAt": snapshot_next_refresh_at(saved_at_epoch, PUBLIC_RATINGS_REFRESH_SECONDS) if saved_at_epoch else "",
            "refreshSeconds": PUBLIC_RATINGS_REFRESH_SECONDS,
            "source": "snapshot",
        },
    }


def read_public_ratings_snapshot(query: dict[str, str] | None = None) -> dict[str, Any] | None:
    snapshot = read_public_ratings_snapshot_file()
    if not snapshot:
        return None
    if snapshot.get("query") != normalize_public_ratings_query(query):
        return None

    payload = snapshot.get("payload")
    if not isinstance(payload, dict):
        return None
    return add_public_snapshot_meta(payload, snapshot)


def read_any_public_ratings_snapshot(query: dict[str, str] | None = None) -> dict[str, Any] | None:
    snapshot = read_public_ratings_snapshot_file()
    if not snapshot:
        return None

    payload = snapshot.get("payload")
    if not isinstance(payload, dict):
        return None

    result = add_public_snapshot_meta(payload, snapshot)
    result["snapshot"]["source"] = "snapshot-fallback"
    result["snapshot"]["queryMismatch"] = snapshot.get("query") != normalize_public_ratings_query(query)
    return result


def refresh_public_ratings_snapshot(query: dict[str, str] | None = None) -> dict[str, Any]:
    normalized_query = normalize_public_ratings_query(query)
    with _PUBLIC_RATINGS_REFRESH_LOCK:
        payload = strip_for_public(resolve_ratings_payload(normalized_query))
        saved_at_epoch = time.time()
        snapshot = {
            "version": 1,
            "query": normalized_query,
            "savedAt": utc_timestamp(),
            "savedAtEpoch": saved_at_epoch,
            "payload": payload,
        }
        write_public_ratings_snapshot_file(snapshot)
        return add_public_snapshot_meta(payload, snapshot)


def start_public_ratings_snapshot_refresh(query: dict[str, str] | None = None) -> bool:
    refresh_key = public_ratings_refresh_key(query)
    with _PUBLIC_RATINGS_REFRESHING_LOCK:
        if refresh_key in _PUBLIC_RATINGS_REFRESHING_KEYS:
            return False
        _PUBLIC_RATINGS_REFRESHING_KEYS.add(refresh_key)

    def refresh() -> None:
        try:
            refresh_public_ratings_snapshot(query)
        except Exception as error:
            sys.stderr.write(f"Public ratings snapshot background refresh failed: {error}\n")
        finally:
            with _PUBLIC_RATINGS_REFRESHING_LOCK:
                _PUBLIC_RATINGS_REFRESHING_KEYS.discard(refresh_key)

    thread = threading.Thread(target=refresh, daemon=True)
    thread.start()
    return True


def public_ratings_snapshot_refreshing() -> bool:
    with _PUBLIC_RATINGS_REFRESHING_LOCK:
        return bool(_PUBLIC_RATINGS_REFRESHING_KEYS)


def public_ratings_snapshot_is_stale(snapshot: dict[str, Any] | None, query: dict[str, str] | None = None) -> bool:
    if not snapshot:
        return True
    if snapshot.get("query") != normalize_public_ratings_query(query):
        return True
    saved_at_epoch = float(snapshot.get("savedAtEpoch") or 0)
    return not saved_at_epoch or time.time() - saved_at_epoch >= PUBLIC_RATINGS_REFRESH_SECONDS


def read_admin_ratings_snapshot_file() -> dict[str, Any] | None:
    if not os.path.exists(ADMIN_RATINGS_SNAPSHOT_PATH):
        return None

    with _ADMIN_RATINGS_SNAPSHOT_LOCK:
        try:
            with open(ADMIN_RATINGS_SNAPSHOT_PATH, "r", encoding="utf-8") as file:
                snapshot = json.load(file)
        except (OSError, json.JSONDecodeError):
            return None

    return snapshot if isinstance(snapshot, dict) else None


def write_admin_ratings_snapshot_file(snapshot: dict[str, Any]) -> None:
    directory = os.path.dirname(ADMIN_RATINGS_SNAPSHOT_PATH)
    if directory:
        os.makedirs(directory, exist_ok=True)
    temp_path = f"{ADMIN_RATINGS_SNAPSHOT_PATH}.tmp"

    with _ADMIN_RATINGS_SNAPSHOT_LOCK:
        with open(temp_path, "w", encoding="utf-8") as file:
            json.dump(snapshot, file, ensure_ascii=False, separators=(",", ":"))
        os.replace(temp_path, ADMIN_RATINGS_SNAPSHOT_PATH)


def add_admin_snapshot_meta(payload: dict[str, Any], snapshot: dict[str, Any]) -> dict[str, Any]:
    saved_at_epoch = float(snapshot.get("savedAtEpoch") or 0)
    return {
        **copy.deepcopy(payload),
        "snapshot": {
            "updatedAt": snapshot.get("savedAt", ""),
            "nextRefreshAt": snapshot_next_refresh_at(saved_at_epoch, ADMIN_RATINGS_REFRESH_SECONDS) if saved_at_epoch else "",
            "refreshSeconds": ADMIN_RATINGS_REFRESH_SECONDS,
            "source": "admin-snapshot",
        },
    }


def read_admin_ratings_snapshot(query: dict[str, str] | None = None) -> dict[str, Any] | None:
    snapshot = read_admin_ratings_snapshot_file()
    if not snapshot:
        return None
    if snapshot.get("query") != normalize_admin_ratings_query(query):
        return None

    payload = snapshot.get("payload")
    if not isinstance(payload, dict):
        return None
    return add_admin_snapshot_meta(payload, snapshot)


def read_any_admin_ratings_snapshot(query: dict[str, str] | None = None) -> dict[str, Any] | None:
    snapshot = read_admin_ratings_snapshot_file()
    if not snapshot:
        return None

    payload = snapshot.get("payload")
    if not isinstance(payload, dict):
        return None

    result = add_admin_snapshot_meta(payload, snapshot)
    result["snapshot"]["source"] = "admin-snapshot-fallback"
    result["snapshot"]["queryMismatch"] = snapshot.get("query") != normalize_admin_ratings_query(query)
    return result


def read_public_from_admin_ratings_snapshot(query: dict[str, str] | None = None) -> dict[str, Any] | None:
    snapshot = read_admin_ratings_snapshot_file()
    if not snapshot:
        return None

    payload = snapshot.get("payload")
    if not isinstance(payload, dict):
        return None

    public_payload = strip_for_public(payload)
    result = add_public_snapshot_meta(public_payload, snapshot)
    result["snapshot"]["source"] = "admin-snapshot-public"
    result["snapshot"]["queryMismatch"] = snapshot.get("query") != normalize_admin_ratings_query(query)
    return result


def refresh_admin_ratings_snapshot(query: dict[str, str] | None = None) -> dict[str, Any]:
    normalized_query = normalize_admin_ratings_query(query)
    with _ADMIN_RATINGS_REFRESH_LOCK:
        payload = resolve_ratings_payload(normalized_query)
        saved_at_epoch = time.time()
        snapshot = {
            "version": 1,
            "query": normalized_query,
            "savedAt": utc_timestamp(),
            "savedAtEpoch": saved_at_epoch,
            "payload": payload,
        }
        write_admin_ratings_snapshot_file(snapshot)
        public_query = normalize_public_ratings_query(normalized_query)
        public_snapshot = {
            "version": 1,
            "query": public_query,
            "savedAt": snapshot["savedAt"],
            "savedAtEpoch": saved_at_epoch,
            "payload": strip_for_public(payload),
        }
        write_public_ratings_snapshot_file(public_snapshot)
        return add_admin_snapshot_meta(payload, snapshot)


def start_admin_ratings_snapshot_refresh(query: dict[str, str] | None = None) -> bool:
    refresh_key = admin_ratings_refresh_key(query)
    with _ADMIN_RATINGS_REFRESHING_LOCK:
        if refresh_key in _ADMIN_RATINGS_REFRESHING_KEYS:
            return False
        _ADMIN_RATINGS_REFRESHING_KEYS.add(refresh_key)

    def refresh() -> None:
        try:
            refresh_admin_ratings_snapshot(query)
        except Exception as error:
            sys.stderr.write(f"Admin ratings snapshot background refresh failed: {error}\n")
        finally:
            with _ADMIN_RATINGS_REFRESHING_LOCK:
                _ADMIN_RATINGS_REFRESHING_KEYS.discard(refresh_key)

    thread = threading.Thread(target=refresh, daemon=True)
    thread.start()
    return True


def admin_ratings_snapshot_is_stale(snapshot: dict[str, Any] | None, query: dict[str, str] | None = None) -> bool:
    if not snapshot:
        return True
    if snapshot.get("query") != normalize_admin_ratings_query(query):
        return True
    saved_at_epoch = float(snapshot.get("savedAtEpoch") or 0)
    return not saved_at_epoch or time.time() - saved_at_epoch >= ADMIN_RATINGS_REFRESH_SECONDS


def public_ratings_snapshot_scheduler() -> None:
    while True:
        try:
            if public_ratings_snapshot_is_stale(read_public_ratings_snapshot_file()):
                start_public_ratings_snapshot_refresh({})
        except Exception as error:
            sys.stderr.write(f"Public ratings snapshot refresh failed: {error}\n")

        time.sleep(max(60, min(PUBLIC_RATINGS_REFRESH_SECONDS, 60 * 60)))


def start_public_ratings_snapshot_scheduler() -> None:
    if PUBLIC_RATINGS_REFRESH_SECONDS <= 0:
        return
    thread = threading.Thread(target=public_ratings_snapshot_scheduler, daemon=True)
    thread.start()


def admin_ratings_snapshot_scheduler() -> None:
    while True:
        try:
            if public_ratings_snapshot_is_stale(read_public_ratings_snapshot_file()):
                start_public_ratings_snapshot_refresh({})
            elif not public_ratings_snapshot_refreshing() and admin_ratings_snapshot_is_stale(read_admin_ratings_snapshot_file()):
                start_admin_ratings_snapshot_refresh({})
        except Exception as error:
            sys.stderr.write(f"Admin ratings snapshot refresh failed: {error}\n")

        time.sleep(max(60, min(ADMIN_RATINGS_REFRESH_SECONDS, 60 * 60)))


def start_admin_ratings_snapshot_scheduler() -> None:
    if ADMIN_RATINGS_REFRESH_SECONDS <= 0:
        return
    thread = threading.Thread(target=admin_ratings_snapshot_scheduler, daemon=True)
    thread.start()


def format_report_number(value: Any, digits: int = 2) -> str:
    return f"{float(value or 0):.{digits}f}".replace(".", ",")


def average_values(values: list[Any]) -> float:
    numbers = [float(value or 0) for value in values]
    return sum(numbers) / len(numbers) if numbers else 0.0


def register_pdf_fonts() -> tuple[str, str]:
    global _PDF_FONT_NAMES
    if _PDF_FONT_NAMES:
        return _PDF_FONT_NAMES

    regular_candidates = (
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/Library/Fonts/DejaVuSans.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
        "/Library/Fonts/Arial.ttf",
    )
    bold_candidates = (
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/Library/Fonts/DejaVuSans-Bold.ttf",
        "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
        "/Library/Fonts/Arial Bold.ttf",
    )
    regular_path = next((path for path in regular_candidates if os.path.exists(path)), "")
    bold_path = next((path for path in bold_candidates if os.path.exists(path)), "")

    if regular_path:
        pdfmetrics.registerFont(TTFont("MarathonSans", regular_path))
        regular_font = "MarathonSans"
    else:
        regular_font = "Helvetica"

    if bold_path:
        pdfmetrics.registerFont(TTFont("MarathonSans-Bold", bold_path))
        bold_font = "MarathonSans-Bold"
    else:
        bold_font = regular_font if regular_path else "Helvetica-Bold"

    _PDF_FONT_NAMES = (regular_font, bold_font)
    return _PDF_FONT_NAMES


def pdf_text(value: Any) -> Paragraph:
    regular_font, _ = register_pdf_fonts()
    style = ParagraphStyle(
        "Cell",
        fontName=regular_font,
        fontSize=7.2,
        leading=8.6,
        textColor=colors.HexColor("#20242a"),
    )
    return Paragraph(xml_escape(normalize_text(value)), style)


def pdf_paragraph(value: Any, style: ParagraphStyle) -> Paragraph:
    text = str(value or "").strip()
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines()]
    return Paragraph(xml_escape("\n".join(lines)).replace("\n", "<br/>"), style)


def find_logo_path() -> str:
    candidates = (
        os.path.join(os.path.dirname(__file__), "logoplanka-cropped.png"),
        os.path.join(os.path.dirname(__file__), "logoplanka.png"),
        os.path.join(os.path.dirname(__file__), "..", "logoplanka.png"),
        os.path.join(os.getcwd(), "logoplanka.png"),
    )
    return next((path for path in candidates if os.path.exists(path)), "")


def build_student_pdf_report(student_name: str, rows: list[dict[str, Any]], period: dict[str, Any]) -> bytes:
    regular_font, bold_font = register_pdf_fonts()
    sorted_rows = sorted(
        rows,
        key=lambda row: (
            str(row.get("subject") or ""),
            str(row.get("level") or ""),
            str(row.get("group") or ""),
        ),
    )

    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    title_white = ParagraphStyle(
        "TitleWhite",
        fontName=bold_font,
        fontSize=19,
        leading=22,
        textColor=colors.white,
    )
    subtitle_white = ParagraphStyle(
        "SubtitleWhite",
        fontName=regular_font,
        fontSize=9.6,
        leading=12,
        textColor=colors.white,
    )
    body_style = ParagraphStyle(
        "ParentBody",
        fontName=regular_font,
        fontSize=9.7,
        leading=12.3,
        textColor=colors.HexColor("#2d3146"),
    )
    body_bold = ParagraphStyle(
        "ParentBodyBold",
        parent=body_style,
        fontName=bold_font,
        fontSize=10.4,
        leading=12.6,
    )
    small_label = ParagraphStyle(
        "SmallLabel",
        fontName=regular_font,
        fontSize=7.4,
        leading=9,
        textColor=colors.HexColor("#8b97ab"),
    )
    student_name_style = ParagraphStyle(
        "StudentName",
        fontName=bold_font,
        fontSize=10.8,
        leading=13,
        textColor=colors.HexColor("#2d3146"),
    )
    section_title = ParagraphStyle(
        "SectionTitle",
        fontName=bold_font,
        fontSize=11.6,
        leading=14,
        textColor=colors.HexColor("#2d3146"),
    )
    table_header_style = ParagraphStyle(
        "TableHeader",
        fontName=bold_font,
        fontSize=7.2,
        leading=8.4,
        alignment=1,
        textColor=colors.white,
    )
    table_cell_style = ParagraphStyle(
        "TableCell",
        fontName=regular_font,
        fontSize=7.3,
        leading=8.6,
        alignment=1,
        textColor=colors.HexColor("#2d3146"),
    )
    table_cell_bold = ParagraphStyle(
        "TableCellBold",
        parent=table_cell_style,
        fontName=bold_font,
    )

    logo_path = find_logo_path()
    if logo_path:
        logo_cell: Any = ReportImage(logo_path, width=18 * mm, height=18 * mm)
    else:
        logo_cell = pdf_paragraph("ПП", ParagraphStyle("LogoFallback", fontName=bold_font, fontSize=12, leading=14, alignment=1, textColor=colors.white))
    logo_box = Table([[logo_cell]], colWidths=[20 * mm], rowHeights=[20 * mm])
    logo_box.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#4562f0")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )

    header = Table(
        [
            [
                logo_box,
                [
                    Paragraph("ОТЧЕТ ПО МАРАФОНАМ", title_white),
                    Paragraph("Ежедневная отработка первой части ЕГЭ/ОГЭ", subtitle_white),
                ],
            ]
        ],
        colWidths=[28 * mm, 153 * mm],
        rowHeights=[24 * mm],
    )
    header.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#4562f0")),
                ("ALIGN", (0, 0), (0, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (0, 0), 3 * mm),
                ("RIGHTPADDING", (0, 0), (0, 0), 1 * mm),
                ("TOPPADDING", (0, 0), (0, 0), 2 * mm),
                ("BOTTOMPADDING", (0, 0), (0, 0), 2 * mm),
                ("LEFTPADDING", (1, 0), (1, 0), 3 * mm),
                ("RIGHTPADDING", (1, 0), (1, 0), 5 * mm),
            ]
        )
    )
    story: list[Any] = [header, Spacer(1, 4 * mm)]

    about = Table(
        [
            [pdf_paragraph("Что такое марафон", body_bold)],
            [
                pdf_paragraph(
                    "Марафон - это ежедневные 30-минутные подборки заданий первой части ЕГЭ/ОГЭ с "
                    "автоматической проверкой. Все задания взяты из банка ФИПИ и помогают системно "
                    "закрыть пробелы и повысить итоговый балл.",
                    body_style,
                )
            ],
        ],
        colWidths=[181 * mm],
    )
    about.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f2f6ff")),
                ("LINEBEFORE", (0, 0), (0, -1), 3, colors.HexColor("#4562f0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 9),
                ("RIGHTPADDING", (0, 0), (-1, -1), 9),
                ("TOPPADDING", (0, 0), (-1, 0), 7),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 2),
                ("TOPPADDING", (0, 1), (-1, 1), 0),
                ("BOTTOMPADDING", (0, 1), (-1, 1), 8),
            ]
        )
    )
    story.extend([about, Spacer(1, 4 * mm)])

    check = Table(
        [[PdfIcon("check"), pdf_paragraph("В формате марафона ученики получают", body_bold)]],
        colWidths=[7 * mm, 167 * mm],
        rowHeights=[7 * mm],
    )
    check.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (0, 0), 0),
                ("RIGHTPADDING", (0, 0), (0, 0), 0),
                ("LEFTPADDING", (1, 0), (1, 0), 7),
            ]
        )
    )
    story.append(check)
    bullet_text = (
        "- ежедневные задания с напоминаниями\n"
        "- еженедельную статистику и рейтинги по группе, предмету и школе\n"
        "- карту ошибок по итогам прохождения\n"
        "- при необходимости разбор сложных заданий (по математике ЕГЭ - с видеоразборами)\n"
        "- возможность задать вопросы преподавателю\n"
        "- систему мотивации с призами за регулярную работу"
    )
    story.extend(
        [
            Spacer(1, 4 * mm),
            pdf_paragraph(bullet_text, body_style),
            Spacer(1, 1.5 * mm),
            pdf_paragraph(
                "Такой формат помогает поддерживать темп подготовки и значительно повышает результат "
                "по первой части экзамена.",
                body_style,
            ),
            Spacer(1, 3 * mm),
            pdf_paragraph("Ученик", small_label),
            pdf_paragraph(student_name, student_name_style),
            Spacer(1, 2.5 * mm),
        ]
    )

    section = Table(
        [[PdfIcon("chart"), pdf_paragraph("Статистика по марафонам", section_title)]],
        colWidths=[7 * mm, 167 * mm],
        rowHeights=[7 * mm],
    )
    section.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (0, 0), 0),
                ("RIGHTPADDING", (0, 0), (0, 0), 0),
                ("LEFTPADDING", (1, 0), (1, 0), 7),
            ]
        )
    )
    story.extend([section, Spacer(1, 3 * mm)])

    table_data: list[list[Any]] = [
        [
            pdf_paragraph("Предмет", table_header_style),
            pdf_paragraph("Дней сделано", table_header_style),
            pdf_paragraph("Дней всего", table_header_style),
            pdf_paragraph("Качество", table_header_style),
            pdf_paragraph("Качество макс", table_header_style),
            pdf_paragraph("Преподаватель", table_header_style),
        ]
    ]
    for row in sorted_rows:
        subject = SUBJECT_LABELS.get(str(row.get("subject") or ""), row.get("subject") or "")
        table_data.append(
            [
                pdf_paragraph(str(subject).lower(), table_cell_bold),
                pdf_paragraph(int(row.get("daysDone") or 0), table_cell_style),
                pdf_paragraph(int(row.get("daysTotal") or 0), table_cell_style),
                pdf_paragraph(format_report_number(row.get("quality"), 0), table_cell_style),
                pdf_paragraph("100", table_cell_style),
                pdf_paragraph(row.get("teacher") or "Без преподавателя", table_cell_style),
            ]
        )

    col_widths = [28 * mm, 30 * mm, 26 * mm, 28 * mm, 30 * mm, 39 * mm]
    report_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    report_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4562f0")),
                ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#d9e5ea")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.extend([report_table, Spacer(1, 5 * mm)])

    parent_note = Table(
        [
            [
                pdf_paragraph(
                    "Со своей стороны мы ежедневно отслеживаем участие ребят, напоминаем о заданиях, "
                    "ведём рейтинги и проводим конкурс с призами для самых активных участников. Но "
                    "практика показывает, что при поддержке родителей результаты марафона становятся "
                    "значительно выше.\n\n"
                    "Будем очень благодарны, если вы сможете уточнять у ребёнка, выполняет ли он "
                    "задания марафона - такая вовлечённость заметно помогает сохранять регулярность и "
                    "повышает эффективность подготовки.",
                    body_style,
                )
            ]
        ],
        colWidths=[181 * mm],
    )
    parent_note.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fff7ed")),
                ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#ffbd3e")),
                ("LEFTPADDING", (0, 0), (-1, -1), 9),
                ("RIGHTPADDING", (0, 0), (-1, -1), 9),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(parent_note)
    document.build(story)
    return buffer.getvalue()


def build_student_telegram_report(student_name: str, rows: list[dict[str, Any]], period: dict[str, Any]) -> str:
    sorted_rows = sorted(
        rows,
        key=lambda row: (
            str(row.get("subject") or ""),
            str(row.get("level") or ""),
            str(row.get("group") or ""),
        ),
    )
    period_label = f"{period.get('from', '...')} — {period.get('to', '...')}"
    average_coefficient = average_values([row.get("coefficient") for row in sorted_rows])
    average_quality = average_values([row.get("quality") for row in sorted_rows])
    average_final = average_values([row.get("finalScore") for row in sorted_rows])
    total_penalty = sum(float(row.get("penalty") or 0) for row in sorted_rows)

    lines = [
        f"Отчёт по марафону: {student_name}",
        f"Период: {period_label}",
        "",
        f"Коэффициент: {format_report_number(average_coefficient)}",
        f"Качество: {format_report_number(average_quality)}",
        f"Штраф: {format_report_number(total_penalty, 0)}",
        f"Итоговый балл: {format_report_number(average_final)}",
        "",
    ]

    for row in sorted_rows:
        subject = SUBJECT_LABELS.get(str(row.get("subject") or ""), row.get("subject") or "")
        lines.extend([
            f"{subject} {row.get('level') or ''} · {row.get('group') or ''}",
            f"Дни: {int(row.get('daysDone') or 0)}/{int(row.get('daysTotal') or 0)} · "
            f"Качество: {format_report_number(row.get('quality'))} · "
            f"Балл: {format_report_number(row.get('baseScore'))} · "
            f"Штраф: {format_report_number(row.get('penalty'), 0)} · "
            f"Итог: {format_report_number(row.get('finalScore'))}",
            f"Место в группе: {int(row.get('groupPlace') or 0)} · "
            f"Место в школе: {int(row.get('schoolPlace') or 0)}",
            f"Преподаватель: {row.get('teacher') or 'Без преподавателя'}",
            "",
        ])

    text = "\n".join(lines).strip()
    if len(text) > 3900:
        return text[:3850].rstrip() + "\n\nОтчёт сокращён, полная версия доступна в админке."
    return text


def report_filename(student_name: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-zА-Яа-яЁё_-]+", "_", normalize_text(student_name)).strip("_")
    return f"marathon_report_{cleaned or 'student'}.pdf"


def send_telegram_message(chat_id: str, text: str) -> dict[str, Any]:
    token = clean_authorization_header(os.getenv("TELEGRAM_BOT_TOKEN", ""))
    if not token:
        raise BackendError("TELEGRAM_BOT_TOKEN is not configured", HTTPStatus.INTERNAL_SERVER_ERROR)
    response = requests.post(
        f"{TELEGRAM_API_BASE}/bot{token}/sendMessage",
        json={
            "chat_id": chat_id,
            "text": text,
            "disable_web_page_preview": True,
        },
        timeout=30,
    )
    if response.status_code >= 400:
        raise BackendError(f"Telegram API error {response.status_code}: {response.text[:500]}", response.status_code)
    return response.json()


def send_telegram_document(chat_id: str, pdf_bytes: bytes, filename: str, caption: str) -> dict[str, Any]:
    token = clean_authorization_header(os.getenv("TELEGRAM_BOT_TOKEN", ""))
    if not token:
        raise BackendError("TELEGRAM_BOT_TOKEN is not configured", HTTPStatus.INTERNAL_SERVER_ERROR)
    response = requests.post(
        f"{TELEGRAM_API_BASE}/bot{token}/sendDocument",
        data={
            "chat_id": chat_id,
            "caption": caption[:1024],
        },
        files={
            "document": (filename, pdf_bytes, "application/pdf"),
        },
        timeout=60,
    )
    if response.status_code >= 400:
        raise BackendError(f"Telegram API error {response.status_code}: {response.text[:500]}", response.status_code)
    return response.json()


def send_telegram_reports(
    rows: list[dict[str, Any]],
    period: dict[str, Any],
    student_name: str = "",
    *,
    send_pdf: bool = False,
) -> dict[str, Any]:
    chat_targets = load_telegram_chat_targets()
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        name = normalize_text(row.get("name"))
        if student_name and name != student_name:
            continue
        if name:
            grouped[name].append(row)

    sent: list[dict[str, Any]] = []
    missing: list[str] = []
    missing_details: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for name in sorted(grouped):
        target = chat_targets.get(normalize_group_name(name)) or {}
        chat_ids = target.get("chatIds") or []
        parents = target.get("parents") or []
        if not chat_ids:
            missing.append(name)
            missing_details.append({"name": name, "reason": "no_chat_id"})
            continue
        pdf_bytes: bytes | None = None
        text = ""
        for chat_id in chat_ids:
            try:
                if send_pdf:
                    if pdf_bytes is None:
                        pdf_bytes = build_student_pdf_report(name, grouped[name], period)
                    caption = f"Отчет по марафону: {name}"
                    result = send_telegram_document(chat_id, pdf_bytes, report_filename(name), caption)
                else:
                    if not text:
                        text = build_student_telegram_report(name, grouped[name], period)
                    result = send_telegram_message(chat_id, text)
                sent.append(
                    {
                        "name": name,
                        "parents": parents,
                        "chatId": chat_id,
                        "messageId": result.get("result", {}).get("message_id"),
                    }
                )
                time.sleep(0.05)
            except Exception as error:
                errors.append({"name": name, "parents": parents, "chatId": chat_id, "error": str(error)})

    return {
        "ok": len(errors) == 0,
        "format": "pdf" if send_pdf else "text",
        "sent": sent,
        "missing": missing,
        "missingDetails": missing_details,
        "errors": errors,
    }


def load_ratings(
    period_from: str,
    period_to: str,
    group_ids: set[int] | None,
    subjects: set[str] | None,
    include_virtual: bool,
    origins: set[str] | None,
    limit: int | None,
) -> dict[str, Any]:
    t_start = time.time()

    groups = selected_groups(
        group_ids=group_ids,
        subjects=subjects,
        include_virtual=include_virtual,
        origins=origins,
    )
    effective_limit = limit if limit is not None else MAX_GROUPS_PER_REQUEST
    if effective_limit > 0 and len(groups) > effective_limit:
        groups = groups[:effective_limit]

    rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    first_attempt_index: dict[tuple[str, str], tuple[datetime, date]] = {}
    first_attempt_stats: dict[str, int] = {"lookups": 0, "matched": 0, "applied": 0, "reducedPenalty": 0}

    t0 = time.time()
    try:
        first_attempt_index = build_first_attempt_index(period_from, period_to)
    except Exception as error:
        errors.append({"source": "firstAttempts", "error": str(error)})
    t_first_attempts_ms = int((time.time() - t0) * 1000)
    attendance_period_from, attendance_period_to = attendance_request_period(period_from, period_to)

    def load_group(group: GroupInfo):
        content = fetch_attendance_xlsx(group.id, attendance_period_from, attendance_period_to)
        return group, parse_attendance_xlsx(content, group, first_attempt_index, first_attempt_stats)

    t1 = time.time()
    with ThreadPoolExecutor(max_workers=max(1, DEFAULT_CONCURRENCY)) as executor:
        futures = {executor.submit(load_group, group): group for group in groups}
        for future in as_completed(futures):
            group = futures[future]
            try:
                _, group_rows = future.result()
                rows.extend(group_rows)
            except Exception as error:
                errors.append({"groupId": group.id, "group": group.name, "error": str(error)})
    t_xlsx_ms = int((time.time() - t1) * 1000)

    rows = dedupe_rating_rows(rows)
    add_places(rows)

    t_total_ms = int((time.time() - t_start) * 1000)
    print(f"[timings] load_ratings: first_attempts={t_first_attempts_ms}ms xlsx={t_xlsx_ms}ms groups={len(groups)} total={t_total_ms}ms", flush=True, file=sys.stderr)

    return {
        "ok": True,
        "period": {"from": period_from, "to": period_to},
        "attendancePeriod": {"from": attendance_period_from, "to": attendance_period_to},
        "groups": [group.__dict__ for group in groups],
        "rows": rows,
        "errors": errors,
        "firstAttempts": {
            "enabled": FIRST_ATTEMPTS_ENABLED,
            "loaded": len(first_attempt_index),
            **first_attempt_stats,
        },
        "timings": {
            "firstAttemptsMs": t_first_attempts_ms,
            "xlsxMs": t_xlsx_ms,
            "totalMs": t_total_ms,
        },
    }


def resolve_ratings_payload(query: dict[str, str]) -> dict[str, Any]:
    default_from, default_to = current_marathon_period()
    config = load_group_config()
    configured_ids, missing_names, missing_candidates = resolve_config_group_ids(fetch_group_tree(), config)
    period_from = query.get("periodFrom") or os.getenv("SOHOLMS_PERIOD_FROM") or default_from
    period_to = query.get("periodTo") or os.getenv("SOHOLMS_PERIOD_TO") or default_to
    group_ids = parse_int_set(query.get("groupIds", "")) or configured_ids or None
    subjects = parse_str_set(query.get("subjects", ""))
    include_virtual = query.get("includeVirtual") == "1" or bool(config.get("includeVirtual"))
    origins = parse_origin_set(query.get("origins", ""))
    limit = parse_limit(query.get("limit", ""))
    cache_key = f"ratings:{period_from}:{period_to}:{group_ids}:{subjects}:{include_virtual}:{origins}:{limit}"
    payload = copy.deepcopy(cached(
        cache_key,
        DEFAULT_CACHE_SECONDS,
        lambda: load_ratings(period_from, period_to, group_ids, subjects, include_virtual, origins, limit),
    ))
    payload = apply_penalty_overrides(payload)
    if missing_names:
        payload = {
            **payload,
            "missingConfigNames": missing_names,
            "missingConfigCandidates": missing_candidates,
        }
    return payload


def parse_int_set(value: str) -> set[int] | None:
    if not value:
        return None
    result = {int(part) for part in re.split(r"[,\s]+", value.strip()) if part}
    return result or None


def parse_str_set(value: str) -> set[str] | None:
    if not value:
        return None
    result = {infer_subject(part.strip()) for part in value.split(",") if part.strip()}
    result.discard("без предмета")
    return result or None


def parse_origin_set(value: str) -> set[str] | None:
    if not value:
        return None
    allowed = {
        "manual": "ManualGroup",
        "manualgroup": "ManualGroup",
        "auto": "AutoGroup",
        "autogroup": "AutoGroup",
    }
    result = set()
    for part in value.split(","):
        key = part.strip().lower()
        if not key:
            continue
        result.add(allowed.get(key, part.strip()))
    return result or None


def parse_limit(value: str) -> int | None:
    if not value:
        return None
    try:
        return max(0, int(value))
    except ValueError:
        raise BackendError("limit must be a number", HTTPStatus.BAD_REQUEST)


def current_month_range() -> tuple[str, str]:
    today = date.today()
    start = today.replace(day=1)
    if today.month == 12:
        end = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        end = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
    return start.isoformat(), end.isoformat()


def current_marathon_period() -> tuple[str, str]:
    return DEFAULT_PERIOD_FROM, date.today().isoformat()


def json_bytes(payload: Any) -> bytes:
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")


def admin_key_matches(value: str) -> bool:
    if not BACKEND_ADMIN_KEY:
        return False
    return hmac.compare_digest(value.strip().encode("utf-8"), BACKEND_ADMIN_KEY.encode("utf-8"))


class Handler(BaseHTTPRequestHandler):
    server_version = "SoholmsMarathonBackend/1.0"

    def do_OPTIONS(self):
        self.send_response(HTTPStatus.NO_CONTENT)
        self.send_cors_headers()
        self.end_headers()

    def do_POST(self):
        try:
            parsed = urlparse(self.path)
            query = {key: values[-1] for key, values in parse_qs(parsed.query).items()}

            if parsed.path == "/api/telegram/send-report":
                self.require_admin(query)
                body = self.read_json_body()
                payload = resolve_ratings_payload(query)
                result = send_telegram_reports(
                    payload.get("rows", []),
                    payload.get("period", {}),
                    normalize_text(body.get("studentName") if isinstance(body, dict) else ""),
                    send_pdf=bool(isinstance(body, dict) and body.get("format") == "pdf"),
                )
                return self.send_json(result, HTTPStatus.OK if result.get("ok") else HTTPStatus.BAD_GATEWAY)

            if parsed.path == "/api/penalty-override":
                self.require_admin(query)
                body = self.read_json_body()
                return self.send_json(save_penalty_override(body))

            if parsed.path == "/api/settings":
                self.require_admin(query)
                body = self.read_json_body()
                settings = save_app_settings(body)
                return self.send_json({"ok": True, **settings})

            self.send_json({"ok": False, "error": "Not found"}, HTTPStatus.NOT_FOUND)
        except BackendError as error:
            self.send_json({"ok": False, "error": str(error)}, error.status)
        except Exception as error:
            self.send_json({"ok": False, "error": str(error)}, HTTPStatus.INTERNAL_SERVER_ERROR)

    def do_GET(self):
        try:
            parsed = urlparse(self.path)
            query = {key: values[-1] for key, values in parse_qs(parsed.query).items()}

            if parsed.path == "/health":
                return self.send_json({"ok": True})

            if parsed.path == "/api/debug/auth":
                self.require_admin(query)
                return self.send_json({
                    "ok": True,
                    "apiToken": token_fingerprint("SOHOLMS_API_TOKEN"),
                    "excelToken": token_fingerprint("SOHOLMS_EXCEL_TOKEN"),
                    "fallbackToken": token_fingerprint("SOHOLMS_TOKEN"),
                    "cacheItems": len(_CACHE),
                })

            if parsed.path == "/api/cache/clear":
                self.require_admin(query)
                clear_cache()
                return self.send_json({"ok": True, "cacheItems": len(_CACHE)})

            if parsed.path == "/api/settings":
                return self.send_json({"ok": True, **load_app_settings()}, cache_seconds=30)

            if parsed.path == "/api/debug/xlsx":
                self.require_admin(query)
                group_id_value = query.get("groupId", "")
                default_from, default_to = current_marathon_period()
                period_from = query.get("periodFrom") or os.getenv("SOHOLMS_PERIOD_FROM") or default_from
                period_to = query.get("periodTo") or os.getenv("SOHOLMS_PERIOD_TO") or default_to
                try:
                    group_id = int(group_id_value)
                except ValueError:
                    raise BackendError("groupId is required", HTTPStatus.BAD_REQUEST)
                content = fetch_attendance_xlsx(group_id, period_from, period_to)
                return self.send_json({
                    "ok": True,
                    "groupId": group_id,
                    "period": {"from": period_from, "to": period_to},
                    **inspect_attendance_xlsx(content),
                })

            if parsed.path == "/api/groups":
                self.require_admin(query)
                search = query.get("search", "").strip()
                if search:
                    return self.send_json({
                        "ok": True,
                        "groups": search_groups(search),
                    })

                config = load_group_config()
                use_config = query.get("configured") == "1"
                configured_ids, missing_names, missing_candidates = resolve_config_group_ids(fetch_group_tree(), config) if use_config else (None, [], {})
                include_virtual = query.get("includeVirtual") == "1" or (use_config and bool(config.get("includeVirtual")))
                groups = selected_groups(
                    group_ids=parse_int_set(query.get("groupIds", "")) or configured_ids,
                    subjects=parse_str_set(query.get("subjects", "")),
                    include_virtual=include_virtual,
                    origins=parse_origin_set(query.get("origins", "")),
                )
                return self.send_json({
                    "ok": True,
                    "groups": [group.__dict__ for group in groups],
                    "missingConfigNames": missing_names,
                    "missingConfigCandidates": missing_candidates,
                })

            if parsed.path == "/api/ratings/refresh":
                if query.get("admin") == "1":
                    self.require_admin(query)
                    start_admin_ratings_snapshot_refresh(query)
                    payload = read_admin_ratings_snapshot(query) or read_any_admin_ratings_snapshot(query) or {
                        "ok": True,
                        "rows": [],
                        "snapshot": {"source": "admin-warming", "refreshing": True},
                    }
                else:
                    start_public_ratings_snapshot_refresh(query)
                    payload = read_public_ratings_snapshot(query) or read_any_public_ratings_snapshot(query) or {
                        "ok": True,
                        "rows": [],
                        "snapshot": {"source": "warming", "refreshing": True},
                    }
                return self.send_json(payload, cache_seconds=15)

            if parsed.path == "/api/ratings":
                if query.get("public") == "1":
                    snapshot = read_public_ratings_snapshot_file()
                    payload = read_public_ratings_snapshot(query)
                    if query.get("refresh") == "1" or public_ratings_snapshot_is_stale(snapshot, query):
                        start_public_ratings_snapshot_refresh(query)
                    if not payload:
                        payload = read_any_public_ratings_snapshot(query)
                    if not payload:
                        payload = read_public_from_admin_ratings_snapshot(query)
                    if not payload:
                        payload = {
                            "ok": True,
                            "rows": [],
                            "snapshot": {"source": "warming", "refreshing": True},
                        }
                    return self.send_json(payload, cache_seconds=60)

                self.require_admin(query)
                snapshot = read_admin_ratings_snapshot_file()
                payload = read_admin_ratings_snapshot(query)
                if admin_ratings_snapshot_is_stale(snapshot, query):
                    start_admin_ratings_snapshot_refresh(query)
                if not payload:
                    payload = read_any_admin_ratings_snapshot(query)
                if not payload:
                    payload = {
                        "ok": True,
                        "rows": [],
                        "snapshot": {"source": "admin-warming", "refreshing": True},
                    }
                is_warming = not payload.get("rows")
                return self.send_json(payload, cache_seconds=5 if is_warming else 60)

            if parsed.path == "/api/error-map":
                return self.send_json(build_error_map_payload(query), cache_seconds=60)

            if parsed.path == "/api/error-analytics":
                return self.send_json(build_error_analytics_payload(query), cache_seconds=60)

            if parsed.path == "/api/student-year":
                student_name = normalize_text(query.get("name", ""))
                if not student_name:
                    raise BackendError("name is required", HTTPStatus.BAD_REQUEST)
                period_from = query.get("from") or "2025-09-01"
                period_to = query.get("to") or "2026-05-31"
                rows = fetch_year_raw_rows(period_from, period_to)
                result = aggregate_student_year(rows, student_name)
                if result is None:
                    return self.send_json({"ok": False, "error": "student not found"}, HTTPStatus.NOT_FOUND)
                return self.send_json({
                    "ok": True,
                    "student": student_name,
                    "period": {"from": period_from, "to": period_to},
                    **result,
                }, cache_seconds=300)

            if parsed.path == "/api/grades-summary":
                self.require_admin(query)
                period_from = query.get("periodFrom") or "2025-09-01"
                period_to = query.get("periodTo") or "2026-05-31"
                rows = fetch_year_raw_rows(period_from, period_to)
                result = aggregate_all_students(rows)
                return self.send_json({
                    "ok": True,
                    "rows": result,
                    "period": {"from": period_from, "to": period_to},
                }, cache_seconds=300)

            if parsed.path == "/api/student-journal":
                student_name = normalize_text(query.get("name", ""))
                if not student_name:
                    raise BackendError("name is required", HTTPStatus.BAD_REQUEST)
                period_from = query.get("from") or "2025-09-01"
                period_to = query.get("to") or "2026-05-31"
                subject_filter = normalize_text(query.get("subject", "")).lower()
                events = fetch_journal_events(period_from, period_to)
                result = aggregate_student_journal(events, student_name, subject_filter)
                if result is None:
                    return self.send_json({"ok": False, "error": "student not found"}, HTTPStatus.NOT_FOUND)
                return self.send_json({
                    "ok": True,
                    "student": student_name,
                    "period": {"from": period_from, "to": period_to},
                    **result,
                }, cache_seconds=300)

            if parsed.path == "/api/journal-summary":
                self.require_admin(query)
                period_from = query.get("periodFrom") or query.get("from") or "2025-09-01"
                period_to = query.get("periodTo") or query.get("to") or "2026-05-31"
                mode = (query.get("mode") or "students").strip().lower()
                if mode not in ("school", "groups", "students"):
                    raise BackendError("mode must be school|groups|students", HTTPStatus.BAD_REQUEST)
                events = fetch_journal_events(period_from, period_to)
                result = aggregate_school_journal(events, mode)
                return self.send_json({
                    "ok": True,
                    "mode": mode,
                    "period": {"from": period_from, "to": period_to},
                    **result,
                }, cache_seconds=300)

            self.send_json({"ok": False, "error": "Not found"}, HTTPStatus.NOT_FOUND)
        except BackendError as error:
            self.send_json({"ok": False, "error": str(error)}, error.status)
        except Exception as error:
            self.send_json({"ok": False, "error": str(error)}, HTTPStatus.INTERNAL_SERVER_ERROR)

    def send_cors_headers(self):
        self.send_header("Access-Control-Allow-Origin", os.getenv("CORS_ORIGIN", "*"))
        self.send_header("Access-Control-Allow-Methods", "GET,POST,OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "content-type,authorization,x-admin-key")
        self.send_header("Access-Control-Allow-Private-Network", "true")

    def require_admin(self, query: dict[str, str]) -> None:
        if not BACKEND_ADMIN_KEY:
            raise BackendError("Admin key not configured", HTTPStatus.SERVICE_UNAVAILABLE)
        value = self.headers.get("x-admin-key", "") or query.get("adminKey", "")
        if not admin_key_matches(value):
            raise BackendError("Forbidden", HTTPStatus.FORBIDDEN)

    def read_json_body(self) -> dict[str, Any]:
        length = int(self.headers.get("content-length") or 0)
        if length <= 0:
            return {}
        raw = self.rfile.read(length)
        try:
            data = json.loads(raw.decode("utf-8"))
        except json.JSONDecodeError:
            raise BackendError("Invalid JSON body", HTTPStatus.BAD_REQUEST)
        return data if isinstance(data, dict) else {}

    def send_json(self, payload: Any, status: int = HTTPStatus.OK, cache_seconds: int = 0):
        body = json_bytes(payload)
        should_gzip = "gzip" in self.headers.get("Accept-Encoding", "").lower() and len(body) > 1024
        if should_gzip:
            body = gzip.compress(body)

        self.send_response(status)
        self.send_header("content-type", "application/json; charset=utf-8")
        if should_gzip:
            self.send_header("content-encoding", "gzip")
        self.send_header("content-length", str(len(body)))
        if cache_seconds > 0 and status == HTTPStatus.OK:
            self.send_header("cache-control", f"public, max-age={cache_seconds}, stale-while-revalidate={DEFAULT_CACHE_SECONDS}")
        else:
            self.send_header("cache-control", "no-store")
        self.send_cors_headers()
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, fmt: str, *args):
        sys.stderr.write("%s - %s\n" % (self.log_date_time_string(), fmt % args))


def main():
    host = os.getenv("HOST", "127.0.0.1")
    port = int(os.getenv("PORT", "8787"))
    start_public_ratings_snapshot_scheduler()
    start_admin_ratings_snapshot_scheduler()
    server = ThreadingHTTPServer((host, port), Handler)
    print(f"Soholms backend listening on http://{host}:{port}", flush=True)
    server.serve_forever()


if __name__ == "__main__":
    main()
